"""
dossier.py
==========

Builds the GFS results workbook: one Excel file that answers "what would this
strategy actually have done to my money", with enough working shown that a
sceptical reader can check it.

Why a workbook rather than the existing text report
---------------------------------------------------
The text report answers "is this edge real". A dossier answers a different
question - "what happened, year by year, trade by trade, after everything" - and
that question needs rows, not prose. Three things in particular only make sense
in tabular form:

* **The friction decomposition.** Every headline number is quoted three times:
  before cost and tax, after cost, and after cost *and* tax. A single net figure
  hides whether a strategy is mediocre or merely over-traded. The three columns
  are three genuinely separate backtest runs, not one run with percentages
  subtracted afterwards - because costs change which trades are affordable, and
  tax paid in year three is capital that cannot compound in year four.

* **Benchmark-relative risk.** CAGR alone cannot distinguish a strategy that
  earned its return from one that borrowed it from beta. Alpha, beta,
  correlation, tracking error and information ratio are all computed against the
  same daily series, aligned on the same calendar.

* **Rolling windows.** A 14-year CAGR is one sample. Every 3-year and 5-year
  window is a few hundred overlapping samples, and the worst of them is the
  number that decides whether a strategy is actually livable.

Reconciliation
--------------
The Positions, Trades, Tax_Ledger and Equity_Curve sheets all come from the same
run and are meant to tie out:

    starting capital
      + sum(Positions.net_pnl)          (gross P&L less both commission legs)
      - sum(Tax_Ledger.total tax paid)
      + cash interest accrued
      + unrealised P&L on open positions
      = final Equity_Curve value

Cash interest is the one line that is not itemised per row; it is reported in
the summary block so the identity can still be checked.
"""

from __future__ import annotations

import logging
import math
from dataclasses import replace
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd

from core import bars

from . import taxes as tax_mod
from .config import GFSConfig
from .engine import GFSBacktestEngine
from .taxes import TaxConfig

logger = logging.getLogger("gfs.dossier")

TRADING_DAYS = 252

# NIFTY 500 on Yahoo. The reference dossier reports against both the large-cap
# index and the broad index, because a 500-name universe beating the NIFTY 50
# may just be a size tilt rather than skill.
NIFTY_50 = "^NSEI"
NIFTY_500 = "^CRSLDX"


# ── small statistics, kept explicit ─────────────────────────────────────────


def _cagr(first: float, last: float, years: float) -> float:
    if first <= 0 or last <= 0 or years <= 0:
        return float("nan")
    return (last / first) ** (1.0 / years) - 1.0


def _returns(values: Sequence[float]) -> List[float]:
    out: List[float] = []
    for prev, cur in zip(values, values[1:]):
        out.append((cur - prev) / prev if prev else 0.0)
    return out


def _max_drawdown(values: Sequence[float]) -> float:
    peak = float("-inf")
    worst = 0.0
    for v in values:
        peak = max(peak, v)
        if peak > 0:
            worst = min(worst, v / peak - 1.0)
    return worst


def _drawdown_series(values: Sequence[float]) -> List[float]:
    peak = float("-inf")
    out: List[float] = []
    for v in values:
        peak = max(peak, v)
        out.append(v / peak - 1.0 if peak > 0 else 0.0)
    return out


def _mean(xs: Sequence[float]) -> float:
    return sum(xs) / len(xs) if xs else 0.0


def _stdev(xs: Sequence[float]) -> float:
    if len(xs) < 2:
        return 0.0
    m = _mean(xs)
    return math.sqrt(sum((x - m) ** 2 for x in xs) / (len(xs) - 1))


def _annual_vol(rets: Sequence[float]) -> float:
    return _stdev(rets) * math.sqrt(TRADING_DAYS)


def _sharpe(rets: Sequence[float]) -> float:
    sd = _stdev(rets)
    return _mean(rets) / sd * math.sqrt(TRADING_DAYS) if sd > 0 else 0.0


def _sortino(rets: Sequence[float]) -> float:
    """Downside deviation uses the full sample in the denominator, not just the
    losing days. Dividing by the count of negative days instead would reward a
    series simply for having few of them, which is the opposite of the point."""
    if len(rets) < 2:
        return 0.0
    downside = math.sqrt(sum(min(r, 0.0) ** 2 for r in rets) / len(rets))
    return _mean(rets) / downside * math.sqrt(TRADING_DAYS) if downside > 0 else 0.0


def _beta_alpha_corr(
    port: Sequence[float], bench: Sequence[float]
) -> Tuple[float, float, float]:
    """Beta, annualised alpha and correlation from paired daily returns."""
    n = min(len(port), len(bench))
    if n < 2:
        return (float("nan"),) * 3
    p, b = list(port[:n]), list(bench[:n])
    mp, mb = _mean(p), _mean(b)
    cov = sum((x - mp) * (y - mb) for x, y in zip(p, b)) / (n - 1)
    var_b = sum((y - mb) ** 2 for y in b) / (n - 1)
    sd_p, sd_b = _stdev(p), _stdev(b)
    beta = cov / var_b if var_b > 0 else float("nan")
    corr = cov / (sd_p * sd_b) if sd_p > 0 and sd_b > 0 else float("nan")
    # Jensen's alpha at a zero risk-free rate, annualised by compounding the
    # daily intercept - not by multiplying it by 252, which overstates it.
    daily_alpha = mp - beta * mb if beta == beta else float("nan")
    alpha = (1.0 + daily_alpha) ** TRADING_DAYS - 1.0 if daily_alpha == daily_alpha else float("nan")
    return beta, alpha, corr


def _tracking_error(port: Sequence[float], bench: Sequence[float]) -> float:
    n = min(len(port), len(bench))
    if n < 2:
        return float("nan")
    diff = [p - b for p, b in zip(port[:n], bench[:n])]
    return _stdev(diff) * math.sqrt(TRADING_DAYS)


def _information_ratio(port: Sequence[float], bench: Sequence[float]) -> float:
    n = min(len(port), len(bench))
    if n < 2:
        return float("nan")
    diff = [p - b for p, b in zip(port[:n], bench[:n])]
    sd = _stdev(diff)
    return _mean(diff) / sd * math.sqrt(TRADING_DAYS) if sd > 0 else float("nan")


# ── reference series ────────────────────────────────────────────────────────


def _index_series(
    symbol: str, calendar: Sequence[pd.Timestamp], start: date, end: date
) -> Optional[List[float]]:
    """Index closes forward-filled onto the portfolio's own trading calendar.

    Forward-filling rather than dropping non-matching days keeps every series the
    same length as the equity curve, so daily returns stay paired. An index
    holiday that the equity calendar does not share shows up as a zero return,
    which is the honest reading: nothing was marked that day.

    The index is synced into the shared bar store on demand. NIFTY 500 is not
    part of any backtest's normal fetch set, so without this the column would
    silently come back blank on a first run.
    """
    try:
        bars.sync([symbol], start, end)
    except Exception as exc:  # noqa: BLE001 - a missing index must not kill the run
        logger.warning("Could not sync reference index %s: %s", symbol, exc)
    frame = bars.read_symbol(symbol, start, end)
    if frame is None or frame.empty or "Close" not in frame.columns:
        logger.warning("No data for reference index %s - column will be blank", symbol)
        return None
    closes = frame["Close"].copy()
    closes.index = pd.to_datetime(closes.index).tz_localize(None).normalize()
    closes = closes[~closes.index.duplicated(keep="last")].sort_index()
    aligned = closes.reindex(
        pd.DatetimeIndex([pd.Timestamp(ts).normalize() for ts in calendar]),
        method="ffill",
    )
    if aligned.isna().all():
        return None
    return [None if pd.isna(v) else float(v) for v in aligned]


def _rebase(series: Optional[Sequence[Optional[float]]], capital: float) -> List[Optional[float]]:
    if not series:
        return []
    base = next((v for v in series if v is not None and v > 0), None)
    if base is None:
        return [None] * len(series)
    return [None if v is None else capital * v / base for v in series]


def _universe_equal_weight(
    panels: Dict[str, Any], calendar: Sequence[pd.Timestamp], capital: float
) -> List[float]:
    """An equal-weighted, daily-rebalanced basket of the whole universe.

    This is the benchmark that actually matters for a stock picker: beating the
    NIFTY 50 out of a 500-name universe may only mean the universe outperformed
    the index. Beating the equal-weighted universe means the *selection* worked.

    Daily rebalancing is not tradeable, and it is not meant to be - it is the
    return of "owning everything you could have picked", which is the correct
    null hypothesis for a picking rule.
    """
    closes = {}
    for sym, panel in panels.items():
        frame = getattr(panel, "frame", None)
        if frame is None or "Close" not in frame.columns:
            continue
        s = frame["Close"]
        s.index = pd.to_datetime(s.index).tz_localize(None).normalize()
        closes[sym] = s[~s.index.duplicated(keep="last")]
    if not closes:
        return [capital] * len(calendar)

    idx = pd.DatetimeIndex([pd.Timestamp(ts).normalize() for ts in calendar])
    matrix = pd.DataFrame({k: v.reindex(idx, method="ffill") for k, v in closes.items()})
    daily = matrix.pct_change()
    # Guard against a name's first ever print showing as an enormous "return".
    daily = daily.where(matrix.shift().notna())
    basket = daily.mean(axis=1, skipna=True).fillna(0.0)

    out: List[float] = []
    level = capital
    for i, r in enumerate(basket):
        if i > 0:
            level *= 1.0 + float(r)
        out.append(level)
    return out


# ── running the three friction variants ─────────────────────────────────────


class Variant:
    """One completed backtest, plus everything derived from its equity curve."""

    def __init__(self, label: str, engine: GFSBacktestEngine, capital: float):
        self.label = label
        self.engine = engine
        self.capital = capital
        self.curve = list(engine.daily_log)
        self.dates = [pd.Timestamp(row["date"]) for row in self.curve]
        self.equity = [float(row["equity"]) for row in self.curve]
        self.returns = _returns(self.equity)

    @property
    def years(self) -> float:
        if len(self.dates) < 2:
            return 0.0
        return (self.dates[-1] - self.dates[0]).days / 365.25

    @property
    def final(self) -> float:
        return self.equity[-1] if self.equity else self.capital


def _run_variant(
    label: str,
    cfg: GFSConfig,
    panels,
    sector_panel,
    regime_panel,
    qualify,
    calendar,
    tax_config: Optional[TaxConfig],
) -> Variant:
    logger.info("Dossier: running variant '%s'", label)
    engine = GFSBacktestEngine(
        cfg, panels, sector_panel, regime_panel, qualify, calendar,
        tax_config=tax_config,
    )
    engine.run(cfg.start_date, cfg.end_date)
    return Variant(label, engine, cfg.starting_capital)


def run_variants(
    cfg: GFSConfig, prepared, tax_cfg: Optional[TaxConfig] = None
) -> Dict[str, Variant]:
    """Three runs of the same rules under three friction regimes.

    They share cached panels, so the marginal cost of the extra two runs is the
    daily loop only - the expensive indicator pass happens once.
    """
    tax_cfg = tax_cfg or TaxConfig()
    panels, calendar, sector_panel, regime_panel, qualify = prepared.panels_for(cfg)

    frictionless = replace(cfg, commission_pct=0.0, slippage_bps=0.0)
    args = (panels, sector_panel, regime_panel, qualify, calendar)
    return {
        "net": _run_variant("net of cost + tax", cfg, *args, tax_cfg),
        "pre_tax": _run_variant("before tax", cfg, *args, None),
        "raw": _run_variant("before cost & tax", frictionless, *args, None),
        "_calendar": calendar,
        "_panels": panels,
    }


# ── metric assembly ─────────────────────────────────────────────────────────


def _metric_column(
    equity: Sequence[float], rets: Sequence[float], years: float,
    capital: float, bench_rets: Optional[Sequence[float]],
) -> Dict[str, float]:
    final = equity[-1] if equity else capital
    col = {
        "CAGR": _cagr(capital, final, years),
        "Absolute return": final / capital - 1.0 if capital else float("nan"),
        f"Final value of {capital:,.0f}": final,
        "Max drawdown": _max_drawdown(equity),
        "Volatility (annual)": _annual_vol(rets),
        "Sharpe (rf = 0)": _sharpe(rets),
        "Sortino (rf = 0)": _sortino(rets),
    }
    if bench_rets is None:
        col.update({
            "Alpha (annual)": 0.0, "Beta": 1.0, "Correlation": 1.0,
            "Tracking error": 0.0, "Information ratio": 0.0, "Excess CAGR": 0.0,
        })
    else:
        beta, alpha, corr = _beta_alpha_corr(rets, bench_rets)
        col.update({
            "Alpha (annual)": alpha,
            "Beta": beta,
            "Correlation": corr,
            "Tracking error": _tracking_error(rets, bench_rets),
            "Information ratio": _information_ratio(rets, bench_rets),
        })
    return col


def _calendar_year_returns(dates, equity) -> Dict[int, float]:
    """Return per calendar year, measured from the last mark of the prior year.

    A partial first year is reported as-is (from inception to 31 December) and
    flagged by the caller, rather than annualised - annualising two months of a
    good run produces a number that means nothing.
    """
    by_year: Dict[int, Tuple[float, float]] = {}
    prev_close: Optional[float] = None
    prev_year: Optional[int] = None
    for ts, eq in zip(dates, equity):
        y = ts.year
        if y != prev_year:
            start = prev_close if prev_close is not None else eq
            by_year[y] = (start, eq)
        else:
            by_year[y] = (by_year[y][0], eq)
        prev_close, prev_year = eq, y
    return {
        y: (end / start - 1.0) if start else float("nan")
        for y, (start, end) in by_year.items()
    }


def _rolling_windows(dates, equity, years: int) -> List[Tuple[pd.Timestamp, pd.Timestamp, float]]:
    """Every window of exactly ``years`` calendar years, stepped monthly.

    Stepping monthly rather than daily keeps the sheet readable while still
    producing enough overlapping samples for the worst window to be meaningful.
    """
    if len(dates) < 2:
        return []
    series = pd.Series(equity, index=pd.DatetimeIndex(dates))
    out = []
    # One anchor per month-start, so windows are comparable across the sample.
    anchors = series.resample("MS").first().dropna().index
    for start in anchors:
        target = start + pd.DateOffset(years=years)
        if target > series.index[-1]:
            break
        lo = series.index[series.index.searchsorted(start)]
        hi_pos = series.index.searchsorted(target)
        if hi_pos >= len(series.index):
            break
        hi = series.index[hi_pos]
        a, b = float(series.loc[lo]), float(series.loc[hi])
        span = (hi - lo).days / 365.25
        out.append((lo, hi, _cagr(a, b, span)))
    return out


# ── sheet builders ──────────────────────────────────────────────────────────


def _positions_rows(engine: GFSBacktestEngine, tax_cfg: TaxConfig, last_prices) -> List[dict]:
    rows = []
    for t in engine.pf.closed:
        long_term = t.holding_days > tax_cfg.long_term_days
        rows.append({
            "ticker": t.symbol,
            "industry": t.sector,
            "entry_date": t.entry_date,
            "exit_date": t.exit_date,
            "hold_days": t.holding_days,
            "entry_px": t.entry_price,
            "exit_px": t.exit_price,
            "return_pct": t.pnl_pct / 100.0,
            "qty": t.quantity,
            "invested": t.entry_value,
            "gross_pnl": t.gross_pnl,
            "costs": t.total_cost,
            "net_pnl": t.net_pnl,
            "st_gain": 0.0 if long_term else t.net_pnl,
            "lt_gain": t.net_pnl if long_term else 0.0,
            "exit_reason": t.exit_reason,
            "r_multiple": t.r_multiple,
            "entry_rsi_m": t.entry_rsi_m,
            "entry_rsi_w": t.entry_rsi_w,
            "entry_rsi_d": t.entry_rsi_d,
            "exit_rsi_d": t.exit_rsi_d,
            "status": "partial" if t.partial else "closed",
        })
    # Open positions are marked to the final close and shown as unrealised, so
    # the sheet accounts for every rupee that ever left cash rather than
    # quietly dropping whatever was still open on the last day.
    for pos in engine.pf.positions.values():
        px = last_prices(pos.symbol)
        px = float(px) if px is not None else pos.entry_price
        gross = (px - pos.entry_price) * pos.quantity
        rows.append({
            "ticker": pos.symbol,
            "industry": pos.sector,
            "entry_date": pos.entry_date,
            "exit_date": None,
            "hold_days": None,
            "entry_px": pos.entry_price,
            "exit_px": px,
            "return_pct": (px - pos.entry_price) / pos.entry_price if pos.entry_price else None,
            "qty": pos.quantity,
            "invested": pos.entry_price * pos.quantity,
            "gross_pnl": gross,
            "costs": pos.entry_cost,
            "net_pnl": gross - pos.entry_cost,
            "st_gain": 0.0,
            "lt_gain": 0.0,
            "exit_reason": "",
            "r_multiple": None,
            "entry_rsi_m": pos.entry_rsi_m,
            "entry_rsi_w": pos.entry_rsi_w,
            "entry_rsi_d": pos.entry_rsi_d,
            "exit_rsi_d": None,
            "status": "open",
        })
    rows.sort(key=lambda r: (r["entry_date"], r["ticker"]))
    return rows


def _trade_rows(engine: GFSBacktestEngine, tax_cfg: TaxConfig) -> List[dict]:
    rows = []
    for f in engine.pf.fills:
        long_term = (f["holding_days"] or 0) > tax_cfg.long_term_days
        net = float(f["net_pnl"])
        rows.append({
            "seq": f["seq"],
            "date": f["date"],
            "ticker": f["symbol"],
            "industry": f["sector"],
            "side": f["side"],
            "reason": f["reason"],
            "qty": f["quantity"],
            "price": f["price"],
            "value": f["value"],
            "cost": f["cost"],
            "net_pnl": net,
            "st_gain": 0.0 if (long_term or f["side"] == "BUY") else net,
            "lt_gain": net if (long_term and f["side"] == "SELL") else 0.0,
            "hold_days": f["holding_days"],
            "entry_px": f["entry_price"],
            "resistance": f["resistance"],
            "stop_level": f["stop_level"],
            "cash_after": f["cash_after"],
        })
    return rows


def _tax_ledger_rows(engine: GFSBacktestEngine, tax_cfg: TaxConfig) -> List[dict]:
    table = tax_mod.apply_to_trades(engine.pf.closed, tax_cfg, use_recorded_costs=True)
    if table.empty:
        return []
    by_year = tax_mod.capital_gains_by_year(table, tax_cfg)
    if by_year.empty:
        return []
    return [
        {
            "financial year": f"FY{fy}",
            "short-term gain": r["short_term_gain"],
            "short-term loss": r["short_term_loss"],
            "long-term gain": r["long_term_gain"],
            "long-term loss": r["long_term_loss"],
            "loss brought forward": r["loss_brought_forward"],
            "loss carried forward": r["loss_carried_forward"],
            "taxable ST": r["taxable_stcg"],
            "taxable LT": r["taxable_ltcg"],
            "tax on ST": r["tax_on_stcg"],
            "tax on LT": r["tax_on_ltcg"],
            "total tax paid": r["tax"],
        }
        for fy, r in by_year.iterrows()
    ]


# ── orchestration ───────────────────────────────────────────────────────────


def build_dossier(
    cfg: GFSConfig,
    prepared,
    out_path: Path,
    *,
    tax_cfg: Optional[TaxConfig] = None,
    title: Optional[str] = None,
) -> Path:
    """Run the three friction variants and write the workbook."""
    tax_cfg = tax_cfg or cfg.tax or TaxConfig()
    variants = run_variants(cfg, prepared, tax_cfg)
    net, pre_tax, raw = variants["net"], variants["pre_tax"], variants["raw"]
    calendar, panels = variants["_calendar"], variants["_panels"]

    dates = net.dates
    capital = cfg.starting_capital
    span_start, span_end = dates[0].date(), dates[-1].date()

    n50 = _index_series(NIFTY_50, dates, span_start, span_end)
    n500 = _index_series(NIFTY_500, dates, span_start, span_end)
    ew = _universe_equal_weight(panels, dates, capital)

    n50_reb, n500_reb = _rebase(n50, capital), _rebase(n500, capital)
    n50_rets = _returns([v for v in n50]) if n50 else []
    n500_rets = _returns([v for v in n500]) if n500 else []
    ew_rets = _returns(ew)

    sheets = _assemble(
        cfg=cfg, tax_cfg=tax_cfg, title=title,
        net=net, pre_tax=pre_tax, raw=raw,
        dates=dates, n50=n50, n500=n500, ew=ew,
        n50_reb=n50_reb, n500_reb=n500_reb,
        n50_rets=n50_rets, n500_rets=n500_rets, ew_rets=ew_rets,
    )
    _write_workbook(sheets, out_path)
    logger.info("Dossier written to %s", out_path)
    return out_path


def _assemble(*, cfg, tax_cfg, title, net, pre_tax, raw, dates,
              n50, n500, ew, n50_reb, n500_reb,
              n50_rets, n500_rets, ew_rets) -> Dict[str, Any]:
    capital = cfg.starting_capital
    years = net.years
    equity, rets = net.equity, net.returns

    # ── Summary ──
    bench_for_alpha = n50_rets or None
    cols = {
        "Portfolio (net of cost+tax)": _metric_column(
            equity, rets, years, capital, bench_for_alpha),
        "Before tax": _metric_column(
            pre_tax.equity, pre_tax.returns, pre_tax.years, capital, bench_for_alpha),
        "Before cost & tax": _metric_column(
            raw.equity, raw.returns, raw.years, capital, bench_for_alpha),
    }
    if n50_reb:
        cols["NIFTY 50"] = _metric_column(
            [v for v in n50_reb if v is not None], n50_rets, years, capital, None)
    if n500_reb:
        # Measured against the NIFTY 50 like every other column. Only the
        # benchmark itself gets the identity row (beta 1, alpha 0, TE 0).
        cols["NIFTY 500"] = _metric_column(
            [v for v in n500_reb if v is not None], n500_rets, years, capital,
            bench_for_alpha)

    bench_cagr = cols.get("NIFTY 50", {}).get("CAGR", float("nan"))
    for name, col in cols.items():
        if name != "NIFTY 50":
            col["Excess CAGR"] = col["CAGR"] - bench_cagr

    order = [
        "CAGR", "Absolute return", f"Final value of {capital:,.0f}",
        "Max drawdown", "Volatility (annual)", "Sharpe (rf = 0)", "Sortino (rf = 0)",
        "Alpha (annual)", "Beta", "Correlation", "Tracking error",
        "Information ratio", "Excess CAGR",
    ]
    headers = list(cols.keys())

    closed = net.engine.pf.closed
    costs_paid = sum(t.total_cost for t in closed) + sum(
        p.entry_cost for p in net.engine.pf.positions.values())
    tax_paid = sum(x["amount"] for x in net.engine.pf.taxes_paid)
    wins = [t for t in closed if t.net_pnl > 0]
    mean_open = _mean([float(r["open_positions"]) for r in net.curve])
    mean_cash = _mean([
        float(r["cash"]) / float(r["equity"]) for r in net.curve if float(r["equity"]) > 0])
    deployed_days = sum(1 for r in net.curve if float(r["open_positions"]) > 0)

    pct = "percent"
    summary_rows: List[Tuple] = [
        (title or _default_title(cfg), None, None, None, None, None),
        ("Portfolio figures are NET of brokerage, slippage and capital-gains tax. "
         "Index figures are raw index levels.", None, None, None, None, None),
        (None,) * 6,
        ("Metric", *headers, *([None] * (5 - len(headers)))),
    ]
    fmt_map = {
        "CAGR": pct, "Absolute return": pct, "Max drawdown": pct,
        "Volatility (annual)": pct, "Alpha (annual)": pct,
        "Tracking error": pct, "Excess CAGR": pct,
    }
    for key in order:
        summary_rows.append((key, *[cols[h].get(key) for h in headers],
                             *([None] * (5 - len(headers)))))

    summary_rows += [
        (None,) * 6,
        ("Costs and tax actually paid", None, None, None, None, None),
        ("Brokerage + impact paid", costs_paid, None, None, None, None),
        ("Capital-gains tax paid", tax_paid, None, None, None, None),
        ("Total frictions", costs_paid + tax_paid, None, None, None, None),
        ("CAGR given up to frictions",
         cols["Before cost & tax"]["CAGR"] - cols["Portfolio (net of cost+tax)"]["CAGR"],
         None, None, None, None),
        ("Total fills", len(net.engine.pf.fills), None, None, None, None),
        ("Round trips", len(closed), None, None, None, None),
        ("Win rate", len(wins) / len(closed) if closed else None, None, None, None, None),
        ("Avg holding days",
         _mean([t.holding_days for t in closed]) if closed else None, None, None, None, None),
        ("Mean positions open", mean_open, None, None, None, None),
        ("Mean cash %", mean_cash, None, None, None, None),
        ("Days with any position",
         deployed_days / len(net.curve) if net.curve else None, None, None, None, None),
        (None,) * 6,
        ("Configuration", None, None, None, None, None),
    ]
    for k, v in _config_rows(cfg, tax_cfg, dates).items():
        summary_rows.append((k, v, None, None, None, None))

    # ── Equity_Curve ──
    dd = _drawdown_series(equity)
    port_reb = equity  # already in rupees from the same starting capital
    eq_rows = [
        {
            "date": ts, "portfolio (net)": eq, "cash": float(r["cash"]),
            "deployed": float(r["deployed"]), "positions open": int(r["open_positions"]),
            "drawdown": d,
            "NIFTY 50 rebased": n50_reb[i] if n50_reb else None,
            "NIFTY 500 rebased": n500_reb[i] if n500_reb else None,
            "universe equal-wt rebased": ew[i] if ew else None,
            "portfolio rebased": port_reb[i],
        }
        for i, (ts, eq, r, d) in enumerate(zip(dates, equity, net.curve, dd))
    ]

    # ── Daily_Returns_Portfolio ──
    def at(seq, i):
        if not seq or i == 0:
            return None
        return seq[i - 1] if i - 1 < len(seq) else None

    daily_rows = [
        {
            "date": ts,
            "portfolio (net)": at(rets, i),
            "NIFTY 50": at(n50_rets, i),
            "NIFTY 500": at(n500_rets, i),
            "universe equal-wt": at(ew_rets, i),
            "portfolio equity": equity[i],
            "positions open": int(net.curve[i]["open_positions"]),
        }
        for i, ts in enumerate(dates)
    ]

    # ── Yearly_Returns ──
    py = _calendar_year_returns(dates, equity)
    b50 = _calendar_year_returns(dates, [v for v in n50_reb]) if n50_reb else {}
    b500 = _calendar_year_returns(dates, [v for v in n500_reb]) if n500_reb else {}
    bew = _calendar_year_returns(dates, ew) if ew else {}
    yearly_rows = [
        {
            "Year": y, "Portfolio": py[y],
            "NIFTY 50": b50.get(y), "NIFTY 500": b500.get(y), "Universe EW": bew.get(y),
            "vs NIFTY 50": py[y] - b50[y] if y in b50 else None,
            "vs NIFTY 500": py[y] - b500[y] if y in b500 else None,
        }
        for y in sorted(py)
    ]

    # ── Rolling windows ──
    def rolling_sheet(n: int) -> List[dict]:
        p = _rolling_windows(dates, equity, n)
        m50 = dict(((a, b), c) for a, b, c in _rolling_windows(
            dates, [v for v in n50_reb], n)) if n50_reb else {}
        m500 = dict(((a, b), c) for a, b, c in _rolling_windows(
            dates, [v for v in n500_reb], n)) if n500_reb else {}
        out = []
        for a, b, c in p:
            v50, v500 = m50.get((a, b)), m500.get((a, b))
            out.append({
                "window start": a, "window end": b, "Portfolio": c,
                "NIFTY 50": v50, "NIFTY 500": v500,
                "excess vs NIFTY 50": c - v50 if v50 is not None else None,
                "excess vs NIFTY 500": c - v500 if v500 is not None else None,
            })
        return out

    last_prices = net.engine._price_lookup(dates[-1])
    return {
        "summary_rows": summary_rows,
        "summary_formats": fmt_map,
        "equity": eq_rows,
        "positions": _positions_rows(net.engine, tax_cfg, last_prices),
        "trades": _trade_rows(net.engine, tax_cfg),
        "yearly": yearly_rows,
        "rolling3": rolling_sheet(3),
        "rolling5": rolling_sheet(5),
        "daily": daily_rows,
        "tax": _tax_ledger_rows(net.engine, tax_cfg),
    }


def _default_title(cfg: GFSConfig) -> str:
    return (
        f"GFS (Grandfather/Father/Son) | G>={cfg.g_rsi_min:g}, "
        f"F>={cfg.f_rsi_min:g}, S<={cfg.s_rsi_entry:g} | exit RSI {cfg.exit_rsi:g} | "
        f"{cfg.atr_stop_mult:g}x ATR stop | {cfg.max_positions} x "
        f"{cfg.max_position_pct:g}% | regime {cfg.regime_mode} | "
        f"universe {cfg.universe_index}"
    )


def _config_rows(cfg: GFSConfig, tax_cfg: TaxConfig, dates) -> Dict[str, Any]:
    span = (dates[-1] - dates[0]).days / 365.25
    return {
        "universe": cfg.universe_index,
        "benchmark": cfg.benchmark,
        "htf_mode": cfg.htf_mode,
        "g_rsi_min (monthly)": cfg.g_rsi_min,
        "f_rsi_min (weekly)": cfg.f_rsi_min,
        "s_rsi_entry (daily)": cfg.s_rsi_entry,
        "entry_trigger": cfg.entry_trigger,
        "exit_mode": cfg.exit_mode,
        "exit_rsi": cfg.exit_rsi,
        "min_headroom_pct": cfg.min_headroom_pct,
        "stop_mode": cfg.stop_mode,
        "atr_stop_mult": cfg.atr_stop_mult,
        "max_holding_days": cfg.max_holding_days or "none",
        "sizing_mode": cfg.sizing_mode,
        "max_positions": cfg.max_positions,
        "max_position_pct": cfg.max_position_pct,
        "rank_by": cfg.rank_by,
        "regime_mode": cfg.regime_mode,
        "min_breadth_pct": cfg.min_breadth_pct,
        "sector_top_n": cfg.sector_top_n,
        "max_per_sector": cfg.max_per_sector,
        "min_price": cfg.min_price,
        "min_turnover_cr": cfg.min_turnover_cr,
        "max_atr_pct": cfg.max_atr_pct,
        "commission_pct": cfg.commission_pct,
        "slippage_bps": cfg.slippage_bps,
        "cash_yield_pct": cfg.cash_yield_pct,
        "stcg_rate (from 23 Jul 2024)": tax_cfg.stcg_rate_pct / 100.0,
        "stcg_rate (legacy)": tax_cfg.stcg_rate_pct_legacy / 100.0,
        "ltcg_rate": tax_cfg.ltcg_rate_pct / 100.0,
        "ltcg_exempt_per_year": tax_cfg.ltcg_exempt_per_year,
        "long_term_days": tax_cfg.long_term_days,
        "start_capital": cfg.starting_capital,
        "signal_price": "close (fills at next open)",
        "period": f"{dates[0].date()} to {dates[-1].date()}  ({span:.2f} years)",
    }


# ── workbook writer ─────────────────────────────────────────────────────────

_PCT = '0.00%'
_MONEY = '#,##0.00'
_DATE = 'yyyy-mm-dd'
_RATIO = '0.000'

# Column formats by header name. Anything not listed is left general, which is
# the right default for text and counts.
_COLUMN_FORMATS = {
    "date": _DATE, "entry_date": _DATE, "exit_date": _DATE,
    "window start": _DATE, "window end": _DATE,
    "portfolio (net)": _MONEY, "cash": _MONEY, "deployed": _MONEY,
    "drawdown": _PCT,
    "NIFTY 50 rebased": _MONEY, "NIFTY 500 rebased": _MONEY,
    "universe equal-wt rebased": _MONEY, "portfolio rebased": _MONEY,
    "portfolio equity": _MONEY,
    "entry_px": _MONEY, "exit_px": _MONEY, "price": _MONEY, "value": _MONEY,
    "invested": _MONEY, "gross_pnl": _MONEY, "costs": _MONEY, "cost": _MONEY,
    "net_pnl": _MONEY, "st_gain": _MONEY, "lt_gain": _MONEY,
    "cash_after": _MONEY, "resistance": _MONEY, "stop_level": _MONEY,
    "qty": '#,##0.####', "return_pct": _PCT, "r_multiple": _RATIO,
    "entry_rsi_m": '0.0', "entry_rsi_w": '0.0', "entry_rsi_d": '0.0',
    "exit_rsi_d": '0.0',
    "Portfolio": _PCT, "NIFTY 50": _PCT, "NIFTY 500": _PCT,
    "Universe EW": _PCT, "vs NIFTY 50": _PCT, "vs NIFTY 500": _PCT,
    "excess vs NIFTY 50": _PCT, "excess vs NIFTY 500": _PCT,
    "universe equal-wt": _PCT,
    "short-term gain": _MONEY, "short-term loss": _MONEY,
    "long-term gain": _MONEY, "long-term loss": _MONEY,
    "loss brought forward": _MONEY, "loss carried forward": _MONEY,
    "taxable ST": _MONEY, "taxable LT": _MONEY,
    "tax on ST": _MONEY, "tax on LT": _MONEY, "total tax paid": _MONEY,
}

_SHEET_NOTES = {
    "Positions": [
        "Every position: what was bought, when it was closed and why.",
        "st_gain / lt_gain classify the net P&L by holding period for tax. "
        "Rows with status 'open' are marked to the final close and are unrealised.",
    ],
    "Trades": [
        "Every fill that moved cash, in order. A scaled-out position is one "
        "round trip in Positions but several rows here.",
    ],
    "Yearly_Returns": [
        "Calendar-year return. The first and last years are partial and are "
        "reported as-is, not annualised.",
        "Portfolio is net of brokerage, slippage and capital-gains tax.",
    ],
    "Rolling_3Y": [
        "Rolling 3-year CAGR, one window per month-start.",
        "Every window of exactly 3 years that fits inside the test period.",
    ],
    "Rolling_5Y": [
        "Rolling 5-year CAGR, one window per month-start.",
        "Every window of exactly 5 years that fits inside the test period.",
    ],
    "Tax_Ledger": [
        "Capital gains by Indian financial year (April-March), with loss set-off "
        "and eight-year carry-forward applied.",
    ],
}


# Column lists for sheets that can legitimately come back empty: a short
# backtest has no five-year windows, a losing book has no tax rows. Without
# these the sheet would be blank with no header, leaving the reader unable to
# tell an empty result from a broken one.
_FALLBACK_COLUMNS = {
    "Positions": [
        "ticker", "industry", "entry_date", "exit_date", "hold_days", "entry_px",
        "exit_px", "return_pct", "qty", "invested", "gross_pnl", "costs",
        "net_pnl", "st_gain", "lt_gain", "exit_reason", "r_multiple",
        "entry_rsi_m", "entry_rsi_w", "entry_rsi_d", "exit_rsi_d", "status",
    ],
    "Trades": [
        "seq", "date", "ticker", "industry", "side", "reason", "qty", "price",
        "value", "cost", "net_pnl", "st_gain", "lt_gain", "hold_days",
        "entry_px", "resistance", "stop_level", "cash_after",
    ],
    "Yearly_Returns": [
        "Year", "Portfolio", "NIFTY 50", "NIFTY 500", "Universe EW",
        "vs NIFTY 50", "vs NIFTY 500",
    ],
    "Rolling_3Y": [
        "window start", "window end", "Portfolio", "NIFTY 50", "NIFTY 500",
        "excess vs NIFTY 50", "excess vs NIFTY 500",
    ],
    "Tax_Ledger": [
        "financial year", "short-term gain", "short-term loss", "long-term gain",
        "long-term loss", "loss brought forward", "loss carried forward",
        "taxable ST", "taxable LT", "tax on ST", "tax on LT", "total tax paid",
    ],
}
_FALLBACK_COLUMNS["Rolling_5Y"] = _FALLBACK_COLUMNS["Rolling_3Y"]


def _write_workbook(sheets: Dict[str, Any], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)

    def table(name: str, rows: List[dict], columns: Optional[List[str]] = None):
        ws = wb.create_sheet(name)
        notes = _SHEET_NOTES.get(name, [])
        for note in notes:
            ws.append([note])
            ws.cell(ws.max_row, 1).font = Font(italic=True)
        if notes:
            ws.append([])
        cols = columns or (
            list(rows[0].keys()) if rows else _FALLBACK_COLUMNS.get(name, [])
        )
        ws.append(cols)
        # Read the row back rather than predicting it: on a sheet with no notes
        # openpyxl reports max_row == 1 while still empty, so appending lands on
        # row 1, not row 2. Guessing wrong here bolds the first data row and
        # leaves a blank gap under the header.
        header_row = ws.max_row
        for c in range(1, len(cols) + 1):
            ws.cell(header_row, c).font = bold
        for r in rows:
            ws.append([_excel_safe(r.get(c)) for c in cols])
        for i, col in enumerate(cols, start=1):
            fmt = _COLUMN_FORMATS.get(col)
            letter = get_column_letter(i)
            if fmt:
                for cell in ws[letter][header_row:]:
                    cell.number_format = fmt
            ws.column_dimensions[letter].width = max(11, min(26, len(str(col)) + 4))
        ws.freeze_panes = ws.cell(header_row + 1, 1)
        return ws

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    fmts = sheets["summary_formats"]
    for row in sheets["summary_rows"]:
        ws.append([_excel_safe(v) for v in row])
        label = row[0]
        r = ws.max_row
        if label and all(v is None for v in row[1:]):
            ws.cell(r, 1).font = bold  # section heading or title
        if label == "Metric":
            for c in range(1, 7):
                ws.cell(r, c).font = bold
        fmt = fmts.get(label) if isinstance(label, str) else None
        for c in range(2, 7):
            cell = ws.cell(r, c)
            if not isinstance(cell.value, (int, float)):
                continue
            if fmt:
                cell.number_format = _PCT
            elif isinstance(label, str) and (
                "Win rate" in label or "cash %" in label or "Days with" in label
            ):
                cell.number_format = _PCT
            elif isinstance(label, str) and label in (
                "Beta", "Correlation", "Information ratio",
                "Sharpe (rf = 0)", "Sortino (rf = 0)",
            ):
                cell.number_format = _RATIO
            elif abs(cell.value) >= 1000:
                cell.number_format = _MONEY
    ws.column_dimensions["A"].width = 34
    for letter in "BCDEF":
        ws.column_dimensions[letter].width = 24
    ws["A1"].alignment = Alignment(wrap_text=False)
    ws.freeze_panes = "B5"

    table("Equity_Curve", sheets["equity"])
    table("Positions", sheets["positions"])
    table("Trades", sheets["trades"])
    table("Yearly_Returns", sheets["yearly"])
    table("Rolling_3Y", sheets["rolling3"])
    table("Rolling_5Y", sheets["rolling5"])
    table("Daily_Returns_Portfolio", sheets["daily"])
    table("Tax_Ledger", sheets["tax"])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _excel_safe(value):
    """openpyxl cannot store NaN/Inf, pandas scalars or Timestamps directly."""
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        # Bare date, not datetime: these are daily bars, so a 00:00:00 time
        # component is noise that shows up in the cell.
        return value.to_pydatetime().date()
    if isinstance(value, float):
        return None if math.isnan(value) or math.isinf(value) else value
    if hasattr(value, "item"):
        try:
            return _excel_safe(value.item())
        except (ValueError, AttributeError):
            return str(value)
    if isinstance(value, (str, int, bool, date)):
        return value
    return str(value)
