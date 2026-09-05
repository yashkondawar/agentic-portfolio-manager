"""
dossier.py
==========

Builds the multi-sheet Excel **results dossier** for the quarterly-results
strategy — the full accounting of a backtest run: what it earned, what it gave
away to costs and taxes, how it compared to the market, and every fill that got
it there.

Why this exists separately from ``run_backtest.py``: that module answers "did
this configuration work?" and emits terse CSVs for the next research iteration.
This one answers "what would actually have happened to my money?", which is a
different question and needs three things the research loop never bothered with:

1. **Capital-gains tax.** A strategy with a 90-day time stop realises virtually
   everything at short-term rates. Ignoring that overstates the outcome by the
   STCG rate on every rupee of gain, which is not a rounding error.
2. **A fill blotter.** The engine records round trips; a dossier needs the
   chronological ledger of legs, with the cash balance after each one.
3. **Benchmarks worth comparing against.** A single index line is not enough to
   tell skill from beta, so NIFTY 50, NIFTY 500 and an equal-weight basket of
   the strategy's own universe are all carried through.

Honesty notes that are surfaced *inside* the workbook rather than buried here:
the live strategy's Tier-2 LLM conviction gate has no point-in-time equivalent
and is therefore absent from these numbers, and the quarterly fundamentals only
reach back ~13 quarters, which caps the testable window.
"""

from __future__ import annotations

import logging
from dataclasses import asdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

import numpy as np
import pandas as pd

from backtesting.gfs.taxes import TaxConfig, financial_year

logger = logging.getLogger(__name__)

TRADING_DAYS = 252

#: Sheet order in the delivered workbook.
SHEETS = (
    "Summary",
    "Equity_Curve",
    "Positions",
    "Trades",
    "Yearly_Returns",
    "Rolling_3Y",
    "Rolling_5Y",
    "Daily_Returns_Portfolio",
    "Tax_Ledger",
)


# ── Small numeric helpers ────────────────────────────────────────────────────


def _years_between(first: date, last: date) -> float:
    return max((last - first).days / 365.25, 1e-9)


def cagr(series: Sequence[float], first: date, last: date) -> Optional[float]:
    if len(series) < 2 or series[0] <= 0 or series[-1] <= 0:
        return None
    return ((series[-1] / series[0]) ** (1.0 / _years_between(first, last)) - 1.0) * 100.0


def daily_returns(series: Sequence[float]) -> np.ndarray:
    arr = np.asarray(series, dtype=float)
    if arr.size < 2:
        return np.zeros(0)
    prev = arr[:-1]
    with np.errstate(divide="ignore", invalid="ignore"):
        out = np.where(prev > 0, arr[1:] / prev - 1.0, 0.0)
    return np.nan_to_num(out, nan=0.0, posinf=0.0, neginf=0.0)


def max_drawdown(series: Sequence[float]) -> float:
    """Most negative peak-to-trough excursion, in percent."""
    worst = 0.0
    peak = float("-inf")
    for value in series:
        peak = max(peak, value)
        if peak > 0:
            worst = min(worst, (value - peak) / peak * 100.0)
    return worst


def drawdown_series(series: Sequence[float]) -> List[float]:
    out: List[float] = []
    peak = float("-inf")
    for value in series:
        peak = max(peak, value)
        out.append((value - peak) / peak * 100.0 if peak > 0 else 0.0)
    return out


def annual_vol(rets: np.ndarray) -> Optional[float]:
    if rets.size < 2:
        return None
    return float(np.std(rets, ddof=1) * np.sqrt(TRADING_DAYS) * 100.0)


def sharpe(rets: np.ndarray) -> Optional[float]:
    """Risk-free rate is zero here, matching the rest of the repo's metrics."""
    if rets.size < 2:
        return None
    sd = float(np.std(rets, ddof=1))
    if sd <= 0:
        return None
    return float(np.mean(rets) / sd * np.sqrt(TRADING_DAYS))


def sortino(rets: np.ndarray) -> Optional[float]:
    if rets.size < 2:
        return None
    downside = rets[rets < 0]
    if downside.size == 0:
        return None
    dd = float(np.sqrt(np.sum(downside ** 2) / rets.size))
    if dd <= 0:
        return None
    return float(np.mean(rets) / dd * np.sqrt(TRADING_DAYS))


def beta_alpha(
    port: np.ndarray, bench: np.ndarray, port_cagr: Optional[float],
    bench_cagr: Optional[float],
) -> Dict[str, Optional[float]]:
    """Beta, correlation, Jensen's alpha, tracking error and information ratio.

    Alpha is annual and computed off CAGR rather than by annualising a daily
    intercept: over a three-year window the daily intercept is dominated by
    noise, and the CAGR form is what an investor would actually check.
    """
    out: Dict[str, Optional[float]] = {
        "beta": None, "correlation": None, "alpha": None,
        "tracking_error": None, "information_ratio": None,
    }
    n = min(port.size, bench.size)
    if n < 40:
        return out
    p, b = port[-n:], bench[-n:]
    var_b = float(np.var(b, ddof=1))
    if var_b > 0:
        out["beta"] = float(np.cov(p, b, ddof=1)[0][1] / var_b)
    if np.std(p) > 0 and np.std(b) > 0:
        out["correlation"] = float(np.corrcoef(p, b)[0][1])
    if out["beta"] is not None and port_cagr is not None and bench_cagr is not None:
        out["alpha"] = port_cagr - out["beta"] * bench_cagr
    excess = p - b
    te = float(np.std(excess, ddof=1) * np.sqrt(TRADING_DAYS) * 100.0)
    out["tracking_error"] = te
    if te > 0:
        out["information_ratio"] = float(np.mean(excess) * TRADING_DAYS * 100.0 / te)
    return out


# ── Tax ledger ───────────────────────────────────────────────────────────────


def classify_trades(trades: Iterable[Any], tax_cfg: TaxConfig) -> List[Dict[str, Any]]:
    """Split each closed round trip into its short- or long-term gain bucket.

    The engine's ``pnl`` is already net of the per-side commission proxy, which
    the live config documents as covering STT and exchange charges as well as
    slippage. Statutory charges are therefore NOT added again here — doing so
    would bill the same STT twice. The commission is the cost model; this
    function only decides which capital-gains bucket the result lands in.
    """
    rows: List[Dict[str, Any]] = []
    for t in trades:
        exit_date = getattr(t, "exit_date", None)
        if exit_date is None:
            continue
        held = int(getattr(t, "holding_days", 0) or 0)
        net = float(getattr(t, "pnl", 0.0) or 0.0)
        long_term = held > tax_cfg.long_term_days
        rows.append({
            "symbol": getattr(t, "symbol", ""),
            "entry_date": getattr(t, "entry_date", None),
            "exit_date": exit_date,
            "fy": financial_year(exit_date),
            "holding_days": held,
            "net_pnl": net,
            "long_term": long_term,
            "st_gain": 0.0 if long_term else net,
            "lt_gain": net if long_term else 0.0,
        })
    return rows


def build_tax_ledger(
    classified: Sequence[Dict[str, Any]], tax_cfg: TaxConfig
) -> pd.DataFrame:
    """Per-financial-year capital gains, with set-off and carry-forward.

    Indian rules modelled: short-term losses set off against both short- and
    long-term gains; long-term losses only against long-term gains; unabsorbed
    losses carry forward. The STCG rate changed on 23 Jul 2024, so the rate for
    a year is weighted by the gains actually booked either side of that date
    instead of applying one rate to a year that straddles it.
    """
    columns = [
        "financial year", "short-term gain", "short-term loss",
        "long-term gain", "long-term loss", "loss brought forward",
        "loss carried forward", "taxable ST", "taxable LT",
        "tax on ST", "tax on LT", "total tax paid",
    ]
    if not classified:
        return pd.DataFrame(columns=columns)

    frame = pd.DataFrame(classified)
    rows: List[Dict[str, Any]] = []
    st_carry = 0.0
    lt_carry = 0.0

    for fy, group in frame.groupby("fy", sort=True):
        st_leg = group.loc[~group["long_term"]]
        lt_leg = group.loc[group["long_term"]]
        st_gain = float(st_leg.loc[st_leg["net_pnl"] > 0, "net_pnl"].sum())
        st_loss = abs(float(st_leg.loc[st_leg["net_pnl"] < 0, "net_pnl"].sum()))
        lt_gain = float(lt_leg.loc[lt_leg["net_pnl"] > 0, "net_pnl"].sum())
        lt_loss = abs(float(lt_leg.loc[lt_leg["net_pnl"] < 0, "net_pnl"].sum()))
        brought_forward = st_carry + lt_carry

        st_net = st_gain - st_loss - st_carry
        st_carry = 0.0
        if st_net < 0:
            st_carry, st_net = -st_net, 0.0

        lt_net = lt_gain - lt_loss - lt_carry
        lt_carry = 0.0
        if lt_net < 0:
            lt_carry, lt_net = -lt_net, 0.0

        # An unabsorbed short-term loss may still shelter long-term gains.
        if st_carry > 0 and lt_net > 0:
            used = min(st_carry, lt_net)
            lt_net -= used
            st_carry -= used

        gains = st_leg.loc[st_leg["net_pnl"] > 0]
        if float(gains["net_pnl"].sum()) > 0:
            rate = float(
                (gains["exit_date"].map(tax_cfg.stcg_rate_for) * gains["net_pnl"]).sum()
                / gains["net_pnl"].sum()
            )
        else:
            rate = tax_cfg.stcg_rate_for(group["exit_date"].max())

        taxable_lt = max(0.0, lt_net - tax_cfg.ltcg_exempt_per_year)
        tax_st = st_net * rate / 100.0
        tax_lt = taxable_lt * tax_cfg.ltcg_rate_pct / 100.0

        rows.append({
            "financial year": f"FY{fy}",
            "short-term gain": st_gain,
            "short-term loss": st_loss,
            "long-term gain": lt_gain,
            "long-term loss": lt_loss,
            "loss brought forward": brought_forward,
            "loss carried forward": st_carry + lt_carry,
            "taxable ST": st_net,
            "taxable LT": taxable_lt,
            "tax on ST": tax_st,
            "tax on LT": tax_lt,
            "total tax paid": tax_st + tax_lt,
        })
    return pd.DataFrame(rows, columns=columns)


def fy_end(fy_label: str) -> date:
    """``FY2024-25`` → 31 Mar 2025, the day that year's liability crystallises."""
    start_year = int(fy_label.replace("FY", "").split("-")[0])
    return date(start_year + 1, 3, 31)


# ── Benchmark and universe series ────────────────────────────────────────────


def rebased_series(
    frame: Optional[pd.DataFrame], calendar: Sequence[date], base: float
) -> List[Optional[float]]:
    """A price frame forward-filled onto the backtest calendar and rebased.

    Forward-fill (never back-fill) so a day the index did not trade inherits the
    last real print rather than borrowing a future one.
    """
    if frame is None or frame.empty or "Close" not in frame.columns:
        return [None] * len(calendar)
    closes = frame["Close"].copy()
    closes.index = pd.to_datetime(closes.index).normalize()
    aligned = closes.reindex(
        pd.DatetimeIndex([pd.Timestamp(d) for d in calendar]), method="ffill"
    )
    first = next((v for v in aligned.tolist() if pd.notna(v) and v > 0), None)
    if first is None:
        return [None] * len(calendar)
    return [
        float(v) / float(first) * base if pd.notna(v) and v > 0 else None
        for v in aligned.tolist()
    ]


def equal_weight_universe(
    frames: Dict[str, pd.DataFrame], calendar: Sequence[date], base: float
) -> List[Optional[float]]:
    """Daily-rebalanced equal-weight basket of every universe name with data.

    This is the "could I have just bought the whole shortlist?" control. It is
    the harshest of the three benchmarks because it holds the same names the
    strategy picks from, so beating it is evidence of selection skill rather
    than of universe construction.
    """
    if not frames or not calendar:
        return [None] * len(calendar)
    index = pd.DatetimeIndex([pd.Timestamp(d) for d in calendar])
    cols = []
    for df in frames.values():
        if df is None or df.empty or "Close" not in df.columns:
            continue
        s = df["Close"].copy()
        s.index = pd.to_datetime(s.index).normalize()
        s = s[~s.index.duplicated(keep="last")]
        cols.append(s.reindex(index, method="ffill"))
    if not cols:
        return [None] * len(calendar)
    panel = pd.concat(cols, axis=1)
    rets = panel.pct_change().replace([np.inf, -np.inf], np.nan)
    mean_ret = rets.mean(axis=1, skipna=True).fillna(0.0)
    curve = base * (1.0 + mean_ret).cumprod()
    return [float(v) for v in curve.tolist()]


# ── Assembling the workbook payload ──────────────────────────────────────────


def _metric_column(
    curve: Sequence[float], calendar: Sequence[date],
    bench_rets: Optional[np.ndarray] = None, bench_cagr: Optional[float] = None,
) -> Dict[str, Optional[float]]:
    rets = daily_returns(curve)
    col_cagr = cagr(curve, calendar[0], calendar[-1]) if calendar else None
    out: Dict[str, Optional[float]] = {
        "CAGR (%)": col_cagr,
        "Absolute return (%)": (curve[-1] / curve[0] - 1.0) * 100.0 if curve and curve[0] else None,
        "Final value": curve[-1] if curve else None,
        "Max drawdown (%)": max_drawdown(curve),
        "Volatility, annual (%)": annual_vol(rets),
        "Sharpe": sharpe(rets),
        "Sortino": sortino(rets),
    }
    if bench_rets is not None:
        stats = beta_alpha(rets, bench_rets, col_cagr, bench_cagr)
        out["Alpha, annual (%)"] = stats["alpha"]
        out["Beta"] = stats["beta"]
        out["Correlation"] = stats["correlation"]
        out["Tracking error (%)"] = stats["tracking_error"]
        out["Information ratio"] = stats["information_ratio"]
        out["Excess CAGR (%)"] = (
            None if col_cagr is None or bench_cagr is None else col_cagr - bench_cagr
        )
    else:
        for key in ("Alpha, annual (%)", "Beta", "Correlation",
                    "Tracking error (%)", "Information ratio", "Excess CAGR (%)"):
            out[key] = None
    return out


def build_dossier(
    cfg: Any,
    engine: Any,
    prices: Any,
    *,
    metrics: Optional[Dict[str, Any]] = None,
    tax_cfg: Optional[TaxConfig] = None,
    nifty500: Optional[pd.DataFrame] = None,
    notes: Sequence[str] = (),
) -> Dict[str, pd.DataFrame]:
    """Turn a finished backtest into the nine dossier sheets."""
    tax_cfg = tax_cfg or TaxConfig()
    pf = engine.pf
    daily = engine.daily_log or pf.equity_curve
    if not daily:
        raise ValueError("Backtest produced no equity curve — nothing to report.")

    calendar = [date.fromisoformat(row["date"]) for row in daily]
    before_tax = [float(row["equity"]) for row in daily]
    cash = [float(row["cash"]) for row in daily]
    deployed = [float(row["deployed"]) for row in daily]
    open_positions = [int(row["open_positions"]) for row in daily]

    # ── Cost and tax overlays ────────────────────────────────────────────────
    # Costs already paid by day t, so adding them back recovers the frictionless
    # curve; taxes are not yet paid, so they are subtracted as each FY closes.
    cost_by_day: Dict[date, float] = {}
    for fill in pf.fills:
        cost_by_day[fill.day] = cost_by_day.get(fill.day, 0.0) + float(fill.cost)
    cum_costs: List[float] = []
    running = 0.0
    for day in calendar:
        running += cost_by_day.get(day, 0.0)
        cum_costs.append(running)
    before_cost_and_tax = [e + c for e, c in zip(before_tax, cum_costs)]

    classified = classify_trades(pf.closed, tax_cfg)
    ledger = build_tax_ledger(classified, tax_cfg)
    tax_due = [
        (fy_end(str(row["financial year"])), float(row["total tax paid"]))
        for _, row in ledger.iterrows()
    ]
    cum_tax: List[float] = []
    running = 0.0
    for day in calendar:
        running = sum(amount for due, amount in tax_due if due <= day)
        cum_tax.append(running)
    net = [e - t for e, t in zip(before_tax, cum_tax)]

    # ── Benchmarks ───────────────────────────────────────────────────────────
    base = float(cfg.starting_capital)
    n50 = rebased_series(getattr(prices, "benchmark", None), calendar, base)
    n500 = rebased_series(nifty500, calendar, base)
    universe_ew = equal_weight_universe(
        getattr(prices, "frames", {}) or {}, calendar, base
    )

    def _clean(series: Sequence[Optional[float]]) -> List[float]:
        out: List[float] = []
        last = base
        for value in series:
            last = float(value) if value is not None else last
            out.append(last)
        return out

    n50_c, n500_c, ew_c = _clean(n50), _clean(n500), _clean(universe_ew)
    n50_rets = daily_returns(n50_c)
    n50_cagr = cagr(n50_c, calendar[0], calendar[-1])

    # ── Summary ──────────────────────────────────────────────────────────────
    columns = {
        "Portfolio (net of cost+tax)": _metric_column(net, calendar, n50_rets, n50_cagr),
        "Before tax": _metric_column(before_tax, calendar, n50_rets, n50_cagr),
        "Before cost & tax": _metric_column(
            before_cost_and_tax, calendar, n50_rets, n50_cagr
        ),
        "NIFTY 50": _metric_column(n50_c, calendar, n50_rets, n50_cagr),
    }
    metric_order = [
        "CAGR (%)", "Absolute return (%)", "Final value", "Max drawdown (%)",
        "Volatility, annual (%)", "Sharpe", "Sortino", "Alpha, annual (%)",
        "Beta", "Correlation", "Tracking error (%)", "Information ratio",
        "Excess CAGR (%)",
    ]

    total_costs = cum_costs[-1] if cum_costs else 0.0
    total_tax = cum_tax[-1] if cum_tax else 0.0
    tax_accrued = (
        float(ledger["total tax paid"].sum()) - total_tax
        if "total tax paid" in ledger.columns else 0.0
    )
    gross_cagr = cagr(before_cost_and_tax, calendar[0], calendar[-1])
    net_cagr = cagr(net, calendar[0], calendar[-1])

    closed = list(pf.closed)
    wins = [t for t in closed if t.pnl > 0]
    summary_rows = _summary_rows(
        cfg=cfg, metrics=metrics or {}, columns=columns, metric_order=metric_order,
        calendar=calendar, total_costs=total_costs, total_tax=total_tax,
        tax_accrued=tax_accrued,
        gross_cagr=gross_cagr, net_cagr=net_cagr, fills=len(pf.fills),
        closed=closed, wins=wins, open_positions=open_positions,
        cash=cash, before_tax=before_tax, tax_cfg=tax_cfg, notes=notes,
    )

    # ── Sheets ───────────────────────────────────────────────────────────────
    sheets: Dict[str, pd.DataFrame] = {}
    sheets["Summary"] = pd.DataFrame(summary_rows, columns=[
        "", "Portfolio (net of cost+tax)", "Before tax",
        "Before cost & tax", "NIFTY 50",
    ])
    sheets["Equity_Curve"] = pd.DataFrame({
        "date": calendar,
        "portfolio (net)": net,
        "cash": cash,
        "deployed": deployed,
        "positions open": open_positions,
        "drawdown": drawdown_series(net),
        "NIFTY 50 rebased": n50,
        "NIFTY 500 rebased": n500,
        "universe equal-wt rebased": universe_ew,
        "portfolio rebased": [v / net[0] * base if net[0] else None for v in net],
    })
    sheets["Positions"] = _positions_sheet(closed, classified)
    sheets["Trades"] = _trades_sheet(pf.fills, classified)
    sheets["Yearly_Returns"] = _yearly_sheet(calendar, net, n50_c, n500_c, ew_c)
    sheets["Rolling_3Y"] = _rolling_sheet(calendar, net, n50_c, n500_c, years=3)
    sheets["Rolling_5Y"] = _rolling_sheet(calendar, net, n50_c, n500_c, years=5)
    sheets["Daily_Returns_Portfolio"] = pd.DataFrame({
        "date": calendar[1:],
        "portfolio (net)": daily_returns(net),
        "NIFTY 50": n50_rets,
        "NIFTY 500": daily_returns(n500_c),
        "universe equal-wt": daily_returns(ew_c),
        "portfolio equity": net[1:],
        "positions open": open_positions[1:],
    })
    sheets["Tax_Ledger"] = ledger
    return sheets


def _summary_rows(**kw) -> List[List[Any]]:
    """The Summary sheet as a list of rows: label plus up to four value columns."""
    cfg, columns, metric_order = kw["cfg"], kw["columns"], kw["metric_order"]
    calendar, closed, wins = kw["calendar"], kw["closed"], kw["wins"]
    tax_cfg, metrics = kw["tax_cfg"], kw["metrics"]
    order = ["Portfolio (net of cost+tax)", "Before tax", "Before cost & tax", "NIFTY 50"]

    def row(label: str, *values: Any) -> List[Any]:
        return [label, *values, *([None] * (4 - len(values)))]

    rows: List[List[Any]] = [
        row("qtr_results — backtest dossier"),
        row(f"Period: {calendar[0].isoformat()} → {calendar[-1].isoformat()} "
            f"({_years_between(calendar[0], calendar[-1]):.2f} years)"),
        row(""),
        row("Metric", *order),
    ]
    for metric in metric_order:
        rows.append(row(metric, *[columns[c].get(metric) for c in order]))

    rows += [
        row(""),
        row("Frictions"),
        row("Brokerage + impact paid", kw["total_costs"]),
        row("Capital-gains tax paid", kw["total_tax"]),
        row("Capital-gains tax accrued, not yet due", kw["tax_accrued"]),
        row("Total frictions", kw["total_costs"] + kw["total_tax"]),
        row("CAGR given up to frictions (pp)",
            None if kw["gross_cagr"] is None or kw["net_cagr"] is None
            else kw["gross_cagr"] - kw["net_cagr"]),
        row(""),
        row("Trade statistics"),
        row("Total fills", kw["fills"]),
        row("Round trips", len(closed)),
        row("Win rate (%)", len(wins) / len(closed) * 100.0 if closed else None),
        row("Avg holding days",
            sum(t.holding_days for t in closed) / len(closed) if closed else None),
        row("Mean positions open",
            sum(kw["open_positions"]) / len(kw["open_positions"]) if kw["open_positions"] else None),
        row("Mean cash (%)",
            sum(c / e * 100.0 for c, e in zip(kw["cash"], kw["before_tax"]) if e > 0)
            / max(1, len([e for e in kw["before_tax"] if e > 0]))),
        row(""),
        row("Configuration"),
    ]
    for key, value in sorted(asdict(cfg).items()):
        if isinstance(value, Path):
            value = str(value)
        rows.append(row(key, value if not isinstance(value, (list, tuple)) else str(value)))

    rows += [
        row(""),
        row("Tax model"),
        row("STCG rate (%) from 23 Jul 2024", tax_cfg.stcg_rate_pct),
        row("STCG rate (%) before that", tax_cfg.stcg_rate_pct_legacy),
        row("LTCG rate (%)", tax_cfg.ltcg_rate_pct),
        row("LTCG annual exemption", tax_cfg.ltcg_exempt_per_year),
        row("Long-term threshold (days)", tax_cfg.long_term_days),
    ]
    if metrics:
        rows += [row(""), row("Engine metrics")]
        for key in ("num_trades", "profit_factor", "deflated_sharpe",
                    "avg_exposure_pct", "beta"):
            if metrics.get(key) is not None:
                rows.append(row(key, metrics[key]))

    rows += [row(""), row("Notes")]
    for note in kw["notes"]:
        rows.append(row(note))
    return rows


def _positions_sheet(closed: Sequence[Any], classified: Sequence[Dict]) -> pd.DataFrame:
    gains = {
        (c["symbol"], c["entry_date"], c["exit_date"]): c for c in classified
    }
    rows = []
    for t in closed:
        tag = gains.get((t.symbol, t.entry_date, t.exit_date), {})
        invested = t.entry_price * t.quantity
        rows.append({
            "ticker": t.symbol,
            "industry": t.sector,
            "entry_date": t.entry_date,
            "exit_date": t.exit_date,
            "hold_days": t.holding_days,
            "entry_px": t.entry_price,
            "exit_px": t.exit_price,
            "return_pct": t.pnl_pct,
            "qty": t.quantity,
            "invested": invested,
            "gross_pnl": getattr(t, "gross_pnl", t.pnl),
            "costs": getattr(t, "costs", 0.0),
            "net_pnl": t.pnl,
            "st_gain": tag.get("st_gain", 0.0),
            "lt_gain": tag.get("lt_gain", 0.0),
            "exit_reason": t.exit_reason,
            "status": "closed",
        })
    return pd.DataFrame(rows, columns=[
        "ticker", "industry", "entry_date", "exit_date", "hold_days", "entry_px",
        "exit_px", "return_pct", "qty", "invested", "gross_pnl", "costs",
        "net_pnl", "st_gain", "lt_gain", "exit_reason", "status",
    ])


def _trades_sheet(fills: Sequence[Any], classified: Sequence[Dict]) -> pd.DataFrame:
    gains = {
        (c["symbol"], c["exit_date"]): c for c in classified
    }
    rows = []
    for f in fills:
        tag = gains.get((f.symbol, f.day), {}) if f.side == "SELL" else {}
        rows.append({
            "seq": f.seq,
            "date": f.day,
            "ticker": f.symbol,
            "industry": f.sector,
            "side": f.side,
            "reason": f.reason,
            "qty": f.quantity,
            "price": f.price,
            "value": f.value,
            "cost": f.cost,
            "net_pnl": f.net_pnl,
            "st_gain": tag.get("st_gain"),
            "lt_gain": tag.get("lt_gain"),
            "hold_days": f.holding_days,
            "entry_px": f.entry_price,
            "anchor": f.anchor,
            "stop_level": f.stop_level,
            "cash_after": f.cash_after,
        })
    return pd.DataFrame(rows, columns=[
        "seq", "date", "ticker", "industry", "side", "reason", "qty", "price",
        "value", "cost", "net_pnl", "st_gain", "lt_gain", "hold_days",
        "entry_px", "anchor", "stop_level", "cash_after",
    ])


def _yearly_sheet(calendar, port, n50, n500, ew) -> pd.DataFrame:
    """Calendar-year returns. Partial first and last years are marked as such."""
    rows = []
    by_year: Dict[int, List[int]] = {}
    for i, day in enumerate(calendar):
        by_year.setdefault(day.year, []).append(i)
    for year in sorted(by_year):
        idx = by_year[year]
        # Anchor on the last print of the previous year so a full year is
        # measured close-to-close rather than from its own first session.
        start = idx[0] - 1 if idx[0] > 0 else idx[0]
        end = idx[-1]

        def ret(series):
            if series[start] <= 0:
                return None
            return (series[end] / series[start] - 1.0) * 100.0

        p, b50, b500, bew = ret(port), ret(n50), ret(n500), ret(ew)
        partial = year in (calendar[0].year, calendar[-1].year)
        rows.append({
            "Year": f"{year}*" if partial else str(year),
            "Portfolio": p,
            "NIFTY 50": b50,
            "NIFTY 500": b500,
            "Universe EW": bew,
            "vs NIFTY 50": None if p is None or b50 is None else p - b50,
            "vs NIFTY 500": None if p is None or b500 is None else p - b500,
        })
    return pd.DataFrame(rows, columns=[
        "Year", "Portfolio", "NIFTY 50", "NIFTY 500", "Universe EW",
        "vs NIFTY 50", "vs NIFTY 500",
    ])


def _rolling_sheet(calendar, port, n50, n500, *, years: int) -> pd.DataFrame:
    """Annualised return over every rolling window that fully fits the run."""
    columns = ["window start", "window end", "Portfolio", "NIFTY 50",
               "NIFTY 500", "excess vs NIFTY 50"]
    rows = []
    span = timedelta(days=int(round(365.25 * years)))
    seen_months = set()
    for i, day in enumerate(calendar):
        key = (day.year, day.month)
        if key in seen_months:
            continue
        seen_months.add(key)
        target = day + span
        j = next((k for k in range(i + 1, len(calendar)) if calendar[k] >= target), None)
        if j is None:
            continue

        def ann(series):
            if series[i] <= 0 or series[j] <= 0:
                return None
            return ((series[j] / series[i]) ** (1.0 / years) - 1.0) * 100.0

        p, b50 = ann(port), ann(n50)
        rows.append({
            "window start": day,
            "window end": calendar[j],
            "Portfolio": p,
            "NIFTY 50": b50,
            "NIFTY 500": ann(n500),
            "excess vs NIFTY 50": None if p is None or b50 is None else p - b50,
        })
    return pd.DataFrame(rows, columns=columns)


# ── Excel output ─────────────────────────────────────────────────────────────

#: Sheets whose data starts below a title block, matching the reference layout.
TITLE_OFFSET = {"Positions": 2, "Yearly_Returns": 3}

_EMPTY_SHEET_NOTE = {
    "Rolling_3Y": (
        "No 3-year window fits this backtest. See Rolling_5Y for the reason: the "
        "strategy's first tradeable signal is capped by how far back point-in-time "
        "quarterly fundamentals reach, which leaves a window shorter than three "
        "years. The sheet is kept so this workbook matches the reference structure."
    ),
    "Rolling_5Y": (
        "No 5-year window fits this backtest. The strategy's entry signal needs "
        "point-in-time quarterly fundamentals, and the cached screener.in history "
        "reaches back only ~13 quarters. Year-on-year growth additionally needs a "
        "year-ago quarter, so the first evaluable result is the Mar-2024 quarter. "
        "The sheet is kept so this workbook matches the reference structure; it "
        "will populate automatically once deeper fundamentals exist."
    ),
}


def write_workbook(
    path: Path, sheets: Dict[str, pd.DataFrame], *, titles: Optional[Dict[str, str]] = None
) -> Path:
    """Write the sheets to ``path`` in the reference workbook's layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    titles = titles or {}
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)

    for name in SHEETS:
        frame = sheets.get(name)
        ws = wb.create_sheet(name)
        offset = TITLE_OFFSET.get(name, 0)
        if name in titles:
            ws.cell(row=1, column=1, value=titles[name]).font = bold
        if frame is None or frame.empty:
            note = _EMPTY_SHEET_NOTE.get(name, "No rows for this sheet.")
            cell = ws.cell(row=max(1, offset) + 1, column=1, value=note)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions["A"].width = 110
            if frame is not None and len(frame.columns):
                for col, header in enumerate(frame.columns, start=1):
                    ws.cell(row=max(1, offset) + 3, column=col, value=str(header)).font = bold
            continue

        header_row = offset + 1
        is_summary = name == "Summary"
        if not is_summary:
            for col, header in enumerate(frame.columns, start=1):
                ws.cell(row=header_row, column=col, value=str(header)).font = bold
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        start_row = header_row if is_summary else header_row + 1
        for r, (_, record) in enumerate(frame.iterrows(), start=start_row):
            for c, value in enumerate(record.tolist(), start=1):
                cell = ws.cell(row=r, column=c, value=_excel_safe(value))
                if isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = _number_format(name, str(frame.columns[c - 1]))

        if is_summary:
            for r in range(1, ws.max_row + 1):
                first = ws.cell(row=r, column=1).value
                if isinstance(first, str) and first and ws.cell(row=r, column=2).value is None:
                    ws.cell(row=r, column=1).font = bold

        for c in range(1, len(frame.columns) + 1):
            width = max(
                12,
                min(38, max(
                    [len(str(frame.columns[c - 1]))]
                    + [len(str(v)) for v in frame.iloc[:200, c - 1].tolist()]
                ) + 2),
            )
            ws.column_dimensions[get_column_letter(c)].width = width

    wb.save(path)
    logger.info("Dossier written to %s", path)
    return path


def _number_format(sheet: str, column: str) -> str:
    """Pick a display format from the sheet AND the column.

    The column name alone is ambiguous: ``portfolio (net)`` is a rupee balance
    on Equity_Curve but a fractional daily return on Daily_Returns_Portfolio,
    and formatting the latter as money renders a whole sheet of zeros.
    """
    name = column.lower()
    if sheet == "Daily_Returns_Portfolio":
        return "#,##0.00" if "equity" in name else "0.0000"
    if sheet in ("Yearly_Returns", "Rolling_3Y", "Rolling_5Y"):
        return "0.00"
    if any(k in name for k in ("sharpe", "sortino", "beta", "correlation",
                               "ratio", "drawdown", "return_pct")):
        return "0.00"
    return "#,##0.00"


def _excel_safe(value: Any) -> Any:
    """openpyxl accepts only primitives, dates and None.

    ``pd.Timestamp`` subclasses ``datetime``/``date``, so it must be narrowed
    BEFORE the plain date check or every date lands in the sheet as a midnight
    timestamp.
    """
    if value is None or value is pd.NaT:
        return None
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, (np.floating,)):
        value = float(value)
    if isinstance(value, float):
        return None if (np.isnan(value) or np.isinf(value)) else value
    if isinstance(value, (str, int, bool, date)):
        return value
    return str(value)
