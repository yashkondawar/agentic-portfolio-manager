"""Read-back test for research/equity_researcher/tools/export_financials_xlsx.py
— the deterministic (zero-token, openpyxl) Excel export of a ticker's
extracted financials.

Not an importable package under src/, so loaded directly by file path
(mirrors tests/test_research/test_convert_and_statement.py's pattern).
Builds a synthetic workspace/<TICKER>/ directory (comprehensive_statement.json
in the exact shape build_comprehensive_statement.py's Node.to_dict() writes,
plus derived_metrics.json / eps_bridge_check.json / red_flags.json), runs
export_ticker(), then re-opens the .xlsx with openpyxl to assert sheet
presence, tree indentation, and period columns.
"""
from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
ER_TOOLS_DIR = REPO_ROOT / "research" / "equity_researcher" / "tools"
EXPORT_PATH = ER_TOOLS_DIR / "export_financials_xlsx.py"


def _load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


@pytest.fixture(scope="module")
def exporter():
    return _load_module("export_financials_xlsx", EXPORT_PATH)


def node(metric, label, level, values, children=None):
    """values: {period: (value, basis)}"""
    return {
        "label": label,
        "metric": metric,
        "level": level,
        "values": {
            p: {"value": v, "unit": "INR_cr", "basis": b, "fact_id": f"F-{metric}-{p}", "method": "reported"}
            for p, (v, b) in values.items()
        },
        "fact_ids": [f"F-{metric}-{p}" for p in values],
        "orphan_level": False,
        "children": children or [],
    }


@pytest.fixture
def workspace(tmp_path):
    ws = tmp_path / "TESTCO"
    (ws / "state").mkdir(parents=True)
    (ws / "facts").mkdir(parents=True)

    term_loan_maturity = node(
        "borrowings_term_loan_maturity_1_3y", "Term loan maturing 1-3y", 3,
        {"FY2024": (120.0, "consolidated")},
    )
    term_loan = node(
        "borrowings_term_loan", "Term loan", 2,
        {"FY2024": (200.0, "consolidated"), "FY2025": (210.0, "consolidated")},
        children=[term_loan_maturity],
    )
    borrowings = node(
        "borrowings_noncurrent", "Borrowings (non-current)", 1,
        {"FY2024": (300.0, "consolidated"), "FY2025": (340.0, "consolidated")},
        children=[term_loan],
    )
    revenue = node(
        "revenue_from_operations", "Revenue from operations", 1,
        {
            "FY2024": (1000.0, "consolidated"),
            "FY2025": (1200.0, "consolidated"),
            "Q1 FY2026": (320.0, "consolidated"),
        },
    )
    cfo = node(
        "cfo", "Cash flow from operations", 1,
        {"FY2024": (150.0, "consolidated"), "FY2025": (180.0, "consolidated")},
    )

    comp_statement = {
        "income_statement": {"revenue_from_operations": revenue},
        "balance_sheet": {"borrowings_noncurrent": borrowings},
        "cash_flow": {"cfo": cfo},
    }
    (ws / "state" / "comprehensive_statement.json").write_text(
        json.dumps(comp_statement), encoding="utf-8"
    )

    derived_facts = {
        "facts": [
            {"metric": "gross_margin", "period": "FY2024", "value": 40.0, "flags": []},
            {"metric": "gross_margin", "period": "FY2025", "value": 45.0, "flags": []},
            {"metric": "roe", "period": "FY2024", "value": 18.5, "flags": []},
            {"metric": "roe", "period": "FY2023", "value": 99.0, "flags": ["superseded"]},
        ]
    }
    (ws / "facts" / "derived_metrics.json").write_text(json.dumps(derived_facts), encoding="utf-8")

    eps_bridge_check = {
        "eps_growth_20pct": {"status": "PASS", "value": {"FY2025": 20.0}, "threshold": 20.0, "note": "consistent"},
        "dilution_consecutive": {"status": "FAIL", "value": {"longest_consecutive_run": 3}, "threshold": 2, "note": "3 consecutive years"},
        "interest_coverage": {"status": "NA", "value": None, "threshold": 3.0, "note": "insufficient data"},
        "_basis": "consolidated",
        "_periods": ["FY2024", "FY2025"],
    }
    (ws / "state" / "eps_bridge_check.json").write_text(json.dumps(eps_bridge_check), encoding="utf-8")

    red_flags = [
        {"id": "RF-01", "category": "working_capital", "flag": "DSO rising", "status": "confirmed",
         "severity": "medium", "confidence": "high", "threshold": "20%", "owner": "forensic-auditor"},
    ]
    (ws / "state" / "red_flags.json").write_text(json.dumps(red_flags), encoding="utf-8")

    return ws


# --- export_ticker orchestration --------------------------------------------


def test_export_ticker_writes_xlsx_at_default_path(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    assert out_path == workspace / "exports" / "TESTCO_financials.xlsx"
    assert out_path.exists()


def test_export_ticker_respects_custom_out_path(exporter, workspace, tmp_path):
    custom = tmp_path / "custom_dir" / "out.xlsx"
    out_path = exporter.export_ticker(workspace, out_path=custom)
    assert out_path == custom
    assert custom.exists()


# --- sheet presence ----------------------------------------------------------


def test_all_expected_sheets_present(exporter, workspace):
    # ER v2.1 took the workbook from 7 tabs to 15: horizontal (YoY) and vertical
    # (common-size) analysis per statement, an Other_metrics tab for the production /
    # capacity / dividend lines that sit outside the three statements, and a Contents
    # tab that states which basis each sheet used. Contents is inserted at index 0
    # deliberately -- it is the sheet a reader opens first.
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    assert wb.sheetnames == [
        "Contents",
        "IS_tree", "BS_tree", "CF_tree",
        "IS_horizontal", "BS_horizontal", "CF_horizontal",
        "IS_vertical", "BS_vertical", "CF_vertical",
        "Other_metrics",
        "Quarterly", "Ratios", "EPS_Bridge", "RedFlags",
    ]


# --- tree sheets: indentation + period columns ------------------------------


def test_bs_tree_has_three_level_indentation(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["BS_tree"]

    header = [c.value for c in ws[1]]
    assert header[0] == "Line item"
    assert "FY2024" in header and "FY2025" in header

    rows_by_label = {}
    for row in ws.iter_rows(min_row=2):
        label_cell = row[0]
        if label_cell.value:
            rows_by_label[label_cell.value] = label_cell

    assert "Borrowings (non-current)" in rows_by_label
    assert "Term loan" in rows_by_label
    assert "Term loan maturing 1-3y" in rows_by_label

    # level-1 root has indent 0, level-2 child indent 2, level-3 grandchild indent 4
    assert rows_by_label["Borrowings (non-current)"].alignment.indent == 0
    assert rows_by_label["Term loan"].alignment.indent == 2
    assert rows_by_label["Term loan maturing 1-3y"].alignment.indent == 4


def test_bs_tree_values_align_to_period_columns(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["BS_tree"]

    header = [c.value for c in ws[1]]
    fy2024_col = header.index("FY2024") + 1
    fy2025_col = header.index("FY2025") + 1

    for row in ws.iter_rows(min_row=2):
        if row[0].value == "Borrowings (non-current)":
            assert row[fy2024_col - 1].value == 300.0
            assert row[fy2025_col - 1].value == 340.0


def test_is_tree_shows_revenue_row(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["IS_tree"]
    labels = [c[0].value for c in ws.iter_rows(min_row=2) if c[0].value]
    assert "Revenue from operations" in labels


def test_empty_statement_sheet_shows_placeholder_message(exporter, tmp_path):
    ws_dir = tmp_path / "EMPTYCO"
    (ws_dir / "state").mkdir(parents=True)
    (ws_dir / "state" / "comprehensive_statement.json").write_text(
        json.dumps({"income_statement": {}, "balance_sheet": {}, "cash_flow": {}}),
        encoding="utf-8",
    )
    exporter_mod = _load_module("export_financials_xlsx_2", EXPORT_PATH)
    out_path = exporter_mod.export_ticker(ws_dir)
    wb = load_workbook(out_path)
    ws = wb["IS_tree"]
    assert ws["A1"].value is not None and "No data" in ws["A1"].value


# --- Quarterly sheet ----------------------------------------------------------


def test_quarterly_sheet_excludes_fy_only_columns(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["Quarterly"]
    header = [c.value for c in ws[1]]
    assert "Q1 FY2026" in header
    assert "FY2024" not in header  # annual-only period columns excluded


def test_quarterly_sheet_has_statement_and_line_item_columns(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["Quarterly"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Statement"
    assert header[1] == "Line item"
    rows = [(r[0].value, r[1].value) for r in ws.iter_rows(min_row=2) if r[1].value]
    assert ("Income Statement", "Revenue from operations") in rows


# --- Ratios sheet --------------------------------------------------------------


def test_ratios_sheet_lists_known_metrics_by_period(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["Ratios"]
    header = [c.value for c in ws[1]]
    assert "FY2024" in header and "FY2025" in header

    rows = {r[0].value: r for r in ws.iter_rows(min_row=2)}
    assert "Gross margin %" in rows
    gm_row = rows["Gross margin %"]
    fy2024_idx = header.index("FY2024")
    assert gm_row[fy2024_idx].value == 40.0


def test_ratios_sheet_excludes_superseded_facts(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["Ratios"]
    header = [c.value for c in ws[1]]
    # FY2023 only exists on the superseded ROE fact -- must not appear as a column
    assert "FY2023" not in header


def test_ratios_sheet_placeholder_when_no_derived_metrics(exporter, tmp_path):
    ws_dir = tmp_path / "NOMETRICS"
    (ws_dir / "state").mkdir(parents=True)
    (ws_dir / "state" / "comprehensive_statement.json").write_text(
        json.dumps({"income_statement": {}, "balance_sheet": {}, "cash_flow": {}}), encoding="utf-8"
    )
    exporter_mod = _load_module("export_financials_xlsx_3", EXPORT_PATH)
    out_path = exporter_mod.export_ticker(ws_dir)
    wb = load_workbook(out_path)
    ws = wb["Ratios"]
    assert "No derived_metrics.json" in ws["A1"].value


# --- EPS_Bridge sheet -----------------------------------------------------------


def test_eps_bridge_sheet_lists_rule_verdicts(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["EPS_Bridge"]
    header = [c.value for c in ws[1]]
    assert header == ["Rule", "Status", "Value", "Threshold", "Note"]

    rows = {r[0].value: r[1].value for r in ws.iter_rows(min_row=2)}
    assert rows["eps_growth_20pct"] == "PASS"
    assert rows["dilution_consecutive"] == "FAIL"
    assert rows["interest_coverage"] == "NA"
    # metadata keys (_basis / _periods) must not leak in as pseudo-rules
    assert "_basis" not in rows
    assert "_periods" not in rows


def test_eps_bridge_sheet_placeholder_when_missing(exporter, tmp_path):
    ws_dir = tmp_path / "NOBRIDGE"
    (ws_dir / "state").mkdir(parents=True)
    (ws_dir / "state" / "comprehensive_statement.json").write_text(
        json.dumps({"income_statement": {}, "balance_sheet": {}, "cash_flow": {}}), encoding="utf-8"
    )
    exporter_mod = _load_module("export_financials_xlsx_4", EXPORT_PATH)
    out_path = exporter_mod.export_ticker(ws_dir)
    wb = load_workbook(out_path)
    ws = wb["EPS_Bridge"]
    assert "run tools/eps_bridge_check.py" in ws["A2"].value


# --- RedFlags sheet -----------------------------------------------------------


def test_red_flags_sheet_lists_ledger_entries(exporter, workspace):
    out_path = exporter.export_ticker(workspace)
    wb = load_workbook(out_path)
    ws = wb["RedFlags"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Id"
    assert "Flag" in header

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[0] == "RF-01" and r[3] == "confirmed" for r in rows)


def test_red_flags_sheet_placeholder_when_empty(exporter, tmp_path):
    ws_dir = tmp_path / "NOFLAGS"
    (ws_dir / "state").mkdir(parents=True)
    (ws_dir / "state" / "comprehensive_statement.json").write_text(
        json.dumps({"income_statement": {}, "balance_sheet": {}, "cash_flow": {}}), encoding="utf-8"
    )
    exporter_mod = _load_module("export_financials_xlsx_5", EXPORT_PATH)
    out_path = exporter_mod.export_ticker(ws_dir)
    wb = load_workbook(out_path)
    ws = wb["RedFlags"]
    assert ws["A2"].value == "No red_flags.json entries found."


# --- CLI plumbing --------------------------------------------------------------


def test_cli_writes_expected_output_path(exporter, workspace):
    import sys
    argv_backup = sys.argv
    sys.argv = ["export_financials_xlsx.py", str(workspace)]
    try:
        exporter.main()
    finally:
        sys.argv = argv_backup
    assert (workspace / "exports" / "TESTCO_financials.xlsx").exists()
