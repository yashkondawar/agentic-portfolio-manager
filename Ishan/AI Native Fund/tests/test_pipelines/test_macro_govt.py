"""Offline tests for afund.data.macro_govt — GST/e-way-bill xlsx parsing
(synthetic workbooks built in-test via openpyxl), ICI PDF table parsing
(fake pdfplumber-shaped object, no real PDF library needed), and the
MoSPI eSankhyiki MCP get_data response parser (fixture JSON string,
including the confirmed real null-growth_rate edge case). No network."""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from afund.data.macro_govt import (
    MacroGovtPipeline,
    _upsert_series,
    find_ewb_fy_urls,
    find_gst_collection_url,
    parse_ewb_workbook,
    parse_gst_collection_sheet_name,
    parse_gst_collection_workbook,
    parse_ici_pdf_tables,
    parse_iip_mcp_response,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
SCHEMA_PATH = REPO_ROOT / "src" / "afund" / "db" / "schema.sql"


@pytest.fixture()
def conn(tmp_path):
    db_path = tmp_path / "afund_test.db"
    connection = sqlite3.connect(str(db_path))
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys=ON;")
    connection.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    connection.commit()
    yield connection
    connection.close()


# ---------------------------------------------------------------------------
# (a) GST collections
# ---------------------------------------------------------------------------

def test_parse_gst_collection_sheet_name_hyphen_and_underscore():
    assert parse_gst_collection_sheet_name("Apr-24") == "2024-04-01"
    # The one observed live typo: underscore instead of hyphen.
    assert parse_gst_collection_sheet_name("Mar_26") == "2026-03-01"


def test_parse_gst_collection_sheet_name_rejects_unrecognized():
    assert parse_gst_collection_sheet_name("Total") is None
    assert parse_gst_collection_sheet_name("Sheet1") is None
    assert parse_gst_collection_sheet_name("Xyz-24") is None  # not a month abbrev


def _build_gst_workbook(tmp_path: Path) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, net_revenue in [("Apr-24", 191759.82), ("May-24", 144328.95), ("Mar_26", 177813.59)]:
        ws = wb.create_sheet(sheet_name)
        # A handful of filler rows before the target label, mirroring the
        # real workbook's shape (label not at a fixed row index).
        ws.append(["Some Header", None, None])
        ws.append(["Total Gross GST Revenue", 150000.0, 160000.0, 6.7])
        ws.append(["Total Net GST Revenue", 140000.0, net_revenue, 5.1])
        ws.append(["Refunds", 10000.0, 11000.0, 10.0])
    return wb


def test_parse_gst_collection_workbook_matches_label_not_row_index(tmp_path):
    wb = _build_gst_workbook(tmp_path)
    rows = parse_gst_collection_workbook(wb)
    assert rows == [
        ("2024-04-01", 191759.82),
        ("2024-05-01", 144328.95),
        ("2026-03-01", 177813.59),
    ]


def test_find_gst_collection_url_scrapes_real_shape():
    html = '<a href="//tutorial.gst.gov.in/offlineutilities/gst_statistics/Gross_Net_Tax_collection.xlsx">GST collections</a>'
    url = find_gst_collection_url(html)
    assert url == "https://tutorial.gst.gov.in/offlineutilities/gst_statistics/Gross_Net_Tax_collection.xlsx"


def test_find_gst_collection_url_none_when_absent():
    assert find_gst_collection_url("<html><body>nothing</body></html>") is None


# ---------------------------------------------------------------------------
# (b) E-way bills
# ---------------------------------------------------------------------------

def test_find_ewb_fy_urls_scrapes_multiple_fy_files():
    html = """
    <a href="//tutorial.gst.gov.in/.../ewb-data-2018-19.xlsx">2018-19</a>
    <a href="//tutorial.gst.gov.in/.../ewb-data-2025-26.xlsx">2025-26</a>
    <a href="//tutorial.gst.gov.in/.../ewb-data-2025-26.xlsx">dup</a>
    <a href="https://example.com/not-ewb.xlsx">decoy</a>
    """
    urls = find_ewb_fy_urls(html)
    assert urls == [
        "https://tutorial.gst.gov.in/.../ewb-data-2018-19.xlsx",
        "https://tutorial.gst.gov.in/.../ewb-data-2025-26.xlsx",
    ]


def _build_ewb_workbook(header_style: str) -> openpyxl.Workbook:
    """Two states x 2 months, mirroring the real file's column-position
    stability despite header-text drift between FY files."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Apr")
    if header_style == "old":
        ws.append(["State Code", "State Name", "Year", "Month",
                    "s1", "INTRA STATE SUPPLIES", "s1b",
                    "s2", "Inter State Outward Supplies", "s2b",
                    "s3", "Inter State Inward Supplies", "s3b"])
    else:
        ws.append(["State Code", "State Name", "Year", "Month",
                    "s1", "WITHIN-STATE", "s1b",
                    "s2", "OUTGOING TO OTHER STATES", "s2b",
                    "s3", "INCOMING FROM OTHER STATES", "s3b"])
    # state_code must be numeric to be recognized as a data row (not a
    # header/title row) — real file uses numeric state codes.
    ws.append([1, "Delhi", 2024, 4, "x", 100.0, "y", "x", 50.0, "y", "x", 25.0, "y"])
    ws.append([2, "Maharashtra", 2024, 4, "x", 200.0, "y", "x", 75.0, "y", "x", 40.0, "y"])
    ws.append(["Total", None, None, None])  # non-numeric state_code -> skipped
    return wb


def test_parse_ewb_workbook_sums_columns_5_8_11_regardless_of_header_text():
    for style in ("old", "new"):
        wb = _build_ewb_workbook(style)
        totals = parse_ewb_workbook(wb)
        # Delhi: 100+50+25=175, Maharashtra: 200+75+40=315 -> 490 total.
        assert totals == {"2024-04-01": 490.0}, f"failed for header style {style!r}"


def test_parse_ewb_workbook_skips_non_numeric_state_rows():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Sheet1")
    ws.append(["Header row", None, None, None, None, None, None, None, None, None, None, None, None])
    ws.append([None, None, None, None])  # too short / blank
    totals = parse_ewb_workbook(wb)
    assert totals == {}


# ---------------------------------------------------------------------------
# (d) ICI Eight Core Industries
# ---------------------------------------------------------------------------

class _FakePage:
    def __init__(self, tables):
        self._tables = tables

    def extract_tables(self):
        return self._tables


class _FakePdf:
    def __init__(self, pages):
        self.pages = pages


ICI_INDEX_HEADER = [
    "Sector", None, "Coal", "Crude Oil", "Natural\nGas", "Refinery\nProducts",
    "Fertilizers", "Steel", "Cement", "Electricity", "Overall\nIndex",
]
ICI_GROWTH_HEADER = [
    "Sector", None, "Coal", "Crude Oil", "Natural\nGas", "Refinery\nProducts",
    "Fertilizers", "Steel", "Cement", "Electricity", "Overall\nGrowth",
]


def test_parse_ici_pdf_tables_extracts_index_table_only():
    index_table = [
        ICI_INDEX_HEADER,
        ["", "Weight", "10.3", "8.9", "6.9", "28.0", "2.6", "17.9", "5.4", "19.9", "100.0"],
        ["", "Apr-25", "179.3", "74.2", "74.1", "131.7", "112.9", "219.2", "204.5", "215.7", "163.3"],
        ["", "May-25", "189.8", "76.5", "75.9", "143.3", "127.9", "225.3", "209.0", "218.5", "170.2"],
    ]
    growth_table = [
        ICI_GROWTH_HEADER,
        ["", "Apr-25", "-6.6", "-2.7", "-0.9", "-4.5", "-4.2", "4.4", "6.3", "1.7", "1.0"],
    ]
    pdf = _FakePdf([_FakePage([]), _FakePage([]), _FakePage([]), _FakePage([index_table, growth_table])])

    result = parse_ici_pdf_tables(pdf)

    assert set(result.keys()) == {"2025-04-01", "2025-05-01"}
    apr = result["2025-04-01"]
    assert apr["ICI_COAL"] == pytest.approx(179.3)
    assert apr["ICI_OVERALL"] == pytest.approx(163.3)
    assert apr["ICI_ELECTRICITY"] == pytest.approx(215.7)
    # Growth table must NOT be stored (index-level only, per docstring).
    assert "ICI_GROWTH" not in apr


def test_parse_ici_pdf_tables_handles_month_in_column_0():
    # Defensive: pdfplumber sometimes splits/merges cells so the month
    # label lands in column 0 instead of column 1. Needs >=3 rows total
    # (header + weight row + >=1 data row) to pass the table's shape
    # guard, mirroring the real PDF's table structure.
    index_table = [
        ICI_INDEX_HEADER,
        ["Weight", "10.3", "8.9", "6.9", "28.0", "2.6", "17.9", "5.4", "19.9", "100.0"],
        ["Apr-25", "179.3", "74.2", "74.1", "131.7", "112.9", "219.2", "204.5", "215.7", "163.3"],
    ]
    pdf = _FakePdf([_FakePage([index_table])])
    result = parse_ici_pdf_tables(pdf)
    assert "2025-04-01" in result
    assert result["2025-04-01"]["ICI_OVERALL"] == pytest.approx(163.3)


def test_parse_ici_pdf_tables_returns_empty_when_shape_not_found():
    pdf = _FakePdf([_FakePage([["not", "the", "right", "shape"]]), _FakePage([])])
    assert parse_ici_pdf_tables(pdf) == {}


# ---------------------------------------------------------------------------
# (e) IIP via MoSPI eSankhyiki MCP
# ---------------------------------------------------------------------------

def _mcp_data_json(rows: list[dict]) -> str:
    return json.dumps({"meta_data": {"totalRecords": len(rows)}, "data": rows})


def test_parse_iip_mcp_response_basic():
    text = _mcp_data_json([
        {"year": 2026, "month": "March", "index": "173.2", "growth_rate": "4.1"},
        {"year": 2026, "month": "February", "index": "158.8", "growth_rate": "5.1"},
    ])
    rows = parse_iip_mcp_response(text)
    assert rows == [
        ("2026-02-01", 158.8, 5.1),
        ("2026-03-01", 173.2, 4.1),
    ]


def test_parse_iip_mcp_response_keeps_index_when_growth_rate_is_null():
    # Real MoSPI data: April 2021's growth_rate is null (the COVID
    # base-period comparator, April 2020, was near-zero, making YoY
    # genuinely undefined) — the index value must still be preserved,
    # not dropped along with the missing growth figure.
    text = _mcp_data_json([
        {"year": 2021, "month": "April", "index": "126.1", "growth_rate": None},
        {"year": 2021, "month": "March", "index": "145.6", "growth_rate": "24.2"},
    ])
    rows = parse_iip_mcp_response(text)
    by_date = {d: (idx, g) for d, idx, g in rows}
    assert by_date["2021-04-01"] == (126.1, None)
    assert by_date["2021-03-01"] == (145.6, 24.2)


def test_parse_iip_mcp_response_skips_unparseable_month_or_index():
    text = _mcp_data_json([
        {"year": 2026, "month": "Notamonth", "index": "173.2", "growth_rate": "4.1"},  # bad month
        {"year": 2026, "month": "March", "index": "not-a-number", "growth_rate": "4.1"},  # bad index
        {"year": None, "month": "March", "index": "173.2", "growth_rate": "4.1"},  # missing year
    ])
    assert parse_iip_mcp_response(text) == []


def test_parse_iip_mcp_response_malformed_json_returns_empty():
    assert parse_iip_mcp_response("not json") == []
    assert parse_iip_mcp_response("") == []


def test_macro_govt_pipeline_iip_parse_splits_index_and_yoy(conn):
    """End-to-end through MacroGovtPipeline.parse(): IIP_INDEX keeps the
    null-growth month, IIP_YOY correctly omits it."""
    iip_text = _mcp_data_json([
        {"year": 2021, "month": "April", "index": "126.1", "growth_rate": None},
        {"year": 2021, "month": "May", "index": "115.1", "growth_rate": "27.6"},
    ])
    raw = {"errors": {}, "iip_mcp_text": iip_text}
    pipeline = MacroGovtPipeline(conn=conn)
    parsed = pipeline.parse(raw)
    assert ("2021-04-01", 126.1) in parsed["IIP_INDEX"]
    assert ("2021-05-01", 115.1) in parsed["IIP_INDEX"]
    yoy_dates = [d for d, _v in parsed["IIP_YOY"]]
    assert "2021-04-01" not in yoy_dates
    assert "2021-05-01" in yoy_dates


# ---------------------------------------------------------------------------
# Upsert idempotency (shared helper, same ON CONFLICT DO UPDATE pattern as
# macro_fred._upsert_series)
# ---------------------------------------------------------------------------

def test_upsert_series_is_idempotent_and_refreshes_revisions(conn):
    written = _upsert_series(conn, "GST_COLLECTIONS", "GST_GOVT", [("2026-04-01", 210909.48)], "INR_cr", "M")
    assert written == 1
    # Revised value on re-run: must update, not duplicate.
    _upsert_series(conn, "GST_COLLECTIONS", "GST_GOVT", [("2026-04-01", 211000.00)], "INR_cr", "M")
    stored = conn.execute(
        "SELECT date, value, unit, freq FROM macro_series WHERE series_code='GST_COLLECTIONS'"
    ).fetchall()
    assert len(stored) == 1
    assert stored[0]["value"] == pytest.approx(211000.00)
    assert stored[0]["unit"] == "INR_cr"
    assert stored[0]["freq"] == "M"
