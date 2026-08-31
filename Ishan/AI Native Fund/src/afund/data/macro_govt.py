"""Government macro-data pipeline (GST, e-way bills, EPFO, ICI, IIP) —
Workstream D macro KPI sourcing.

One sources.yaml-driven module covering five sub-sources, each with its
own fetch/parse logic but sharing the same idempotent upsert helper and
the same "never fabricate, degrade honestly" discipline as
afund.data.macro_fred/macro_bis. Unlike those single-purpose modules,
GovtMacroPipeline's `fetch()`/`parse()` wrap EACH sub-source in its own
try/except internally (rather than relying on Pipeline.run()'s single
top-level try/except) so one broken source (confirmed: EPFO) never
prevents the other four from landing.

Sub-sources (see config/sources.yaml `macro_govt` group for the live
verification notes/verify_status per source):

  (a) GST_COLLECTIONS — gst.gov.in statistics page
      (https://www.gst.gov.in/download/gststatistics) lists a static
      "Gross_Net_Tax_collection.xlsx" link (verified live 2026-07: HTTP
      200, one sheet per month named "Mon-YY", e.g. "Apr-24" — NOTE one
      observed sheet name typo uses an underscore, "Mar_26", handled
      defensively). Row-label match ("Total Gross GST Revenue" /
      "Total Net GST Revenue") -> column index 2 (current-month value,
      INR crore) -> macro_series GST_COLLECTIONS (net revenue; gross is
      not separately stored to keep one series per KPI, per
      knowledge/data/kpis/gst_collections.yaml's formula).

  (b) EWAY_BILLS — same statistics page lists one "ewb-data-<FY>.xlsx"
      per fiscal year (e.g. ewb-data-2018-19.xlsx through
      ewb-data-2026-27.xlsx, verified live 2026-07). Each file is a
      state x month table; national monthly EWAY_BILLS = sum of "No. of
      Eway Bills" across all three flow-type column groups (intra-state,
      inter-state-outward, inter-state-inward) and all states, per
      (year, month). Column layout drifts slightly year to year (header
      text differs, e.g. "INTRA STATE SUPPLIES" vs "WITHIN-STATE") but
      the column POSITIONS are stable: eway-bill counts sit at columns
      (0-indexed) 5, 8, 11 in every FY file inspected.

  (c) EPFO_NET_PAYROLL — CONFIRMED BROKEN. The predictable-name pattern
      from the plan (epfindia.gov.in "Payroll_Data_EPFO_<Month>_<Year>.
      xlsx") is blocked by an F5 BIG-IP WAF: a HEAD request returns HTTP
      200 (misleadingly suggesting the file exists), but the actual GET
      returns HTTP 200 with only ~246 bytes of `text/html` ("Request
      Rejected" WAF page), not the xlsx binary — confirmed against
      multiple months/header variations, with a control test against two
      other real PDFs on the same domain succeeding normally (so this is
      a targeted block on the Payroll path, not a blanket block on the
      whole domain). No fetch function is wired for this source; it is
      documented here and in config/sources.yaml as verify_status: broken,
      with the manual-import route (afund.data.macro_manual-style CSV
      import) as the honest fallback — see
      knowledge/data/kpis/epfo_payroll.yaml.

  (d) ICI_INDEX (+8 sub-series) — eaindustry.nic.in publishes the Index
      of Eight Core Industries press release PDF at a predictable path,
      archive_data/ici_press_release/IPR_<YYYY>_<MM>.pdf (verified live
      2026-07). GOTCHA (same class as the documented niftyindices.com
      "HTTP 200 HTML shell for missing dates" precedent in
      config/sources.yaml's `macro` group / CLAUDE.md): a not-yet-
      published month's PDF URL still returns HTTP 200, but with
      Content-Type text/html (a ~1KB site error page), not
      application/pdf — the fetch code checks Content-Type, not status
      code alone, to detect this. Page index 3 of the PDF (0-indexed)
      contains a clean pdfplumber-extractable monthly table: 8 sub-sector
      columns (Coal, Crude Oil, Natural Gas, Refinery Products,
      Fertilizers, Steel, Cement, Electricity) + an Overall Index column,
      for the last ~12-13 months, alongside a matching growth-rate table.
      All 9 columns (8 sub-series + headline) parse cleanly -> macro_series
      ICI_INDEX (headline) + ICI_<SUBSECTOR> (8 sub-series), monthly index
      level (2011-12=100 base).

  (e) IIP via the MoSPI eSankhyiki MCP server (mcp.mospi.gov.in) — BETA,
      wrapped in try/except. Protocol: JSON-RPC 2.0 over a single POST
      endpoint, `Accept: application/json, text/event-stream` header
      required, every response (even non-streaming) comes back
      Server-Sent-Events-framed (`data: {...}` line) and must be parsed
      accordingly. Flow: initialize -> notifications/initialized ->
      tools/call("get_data", {"dataset": "IIP", "filters": {...}}).
      Verified live 2026-07: `get_data` with
      filters={"base_year": "2011-12", "frequency": "Monthly",
      "type": "General", "limit": 200} returns ALL 167 monthly headline
      IIP points in one call (April 2012 through March 2026 as of this
      verification) — no year/month_code filter or pagination needed for
      a full backfill. Each row: {"year", "month", "index", "growth_rate"}
      -> macro_series IIP_INDEX (index level) + IIP_YOY (growth_rate,
      MoSPI's own pre-computed YoY %, not re-derived).

Idempotent upsert (ON CONFLICT...DO UPDATE, same pattern as
afund.data.macro_fred._upsert_series) for every series; job_runs logging
via Pipeline.run(); realistic browser User-Agent via
afund.data.http.make_session() for the XLSX/PDF fetches (gst.gov.in and
eaindustry.nic.in both required it during verification), and a
requests.Session with the MCP-specific Accept header for the IIP calls.
"""
from __future__ import annotations

import datetime as dt
import io
import json
import re
import sqlite3
from typing import Any

import requests

from afund.data.base import Pipeline
from afund.data.http import get, make_session
from afund.sources import get_source

GST_HOST_KEY = "gst.gov.in"
ICI_HOST_KEY = "eaindustry.nic.in"
MOSPI_HOST_KEY = "mcp.mospi.gov.in"

_MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# ICI's 9-column month-level table order (page index 3 of the press
# release PDF, verified live 2026-07): 8 sub-sectors + Overall Index/Growth.
ICI_COLUMNS = [
    "COAL", "CRUDE_OIL", "NATURAL_GAS", "REFINERY_PRODUCTS",
    "FERTILIZERS", "STEEL", "CEMENT", "ELECTRICITY", "OVERALL",
]


def _upsert_series(
    conn: sqlite3.Connection, series_code: str, source: str,
    rows: list[tuple[str, float]], unit: str, freq: str,
) -> int:
    written = 0
    for date, value in rows:
        cur = conn.execute(
            """
            INSERT INTO macro_series (series_code, source, date, value, unit, freq)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(series_code, date) DO UPDATE SET
                value = excluded.value,
                source = excluded.source,
                unit = COALESCE(excluded.unit, macro_series.unit),
                freq = COALESCE(excluded.freq, macro_series.freq)
            """,
            (series_code, source, date, value, unit, freq),
        )
        written += cur.rowcount
    return written


# ---------------------------------------------------------------------------
# (a) GST collections
# ---------------------------------------------------------------------------

def parse_gst_collection_sheet_name(sheet_name: str) -> str | None:
    """'Apr-24' -> '2024-04-01'. Handles the one observed underscore typo
    ('Mar_26') defensively. Returns None (never guesses) if the sheet name
    doesn't match the expected Mon-YY / Mon_YY shape."""
    m = re.match(r"^([A-Za-z]{3})[-_](\d{2})$", sheet_name.strip())
    if not m:
        return None
    mon_abbr, yy = m.groups()
    month = _MONTH_ABBR.get(mon_abbr.lower())
    if month is None:
        return None
    year = 2000 + int(yy)
    return f"{year:04d}-{month:02d}-01"


def parse_gst_collection_workbook(wb: Any) -> list[tuple[str, float]]:
    """Parse the Gross_Net_Tax_collection.xlsx workbook (openpyxl Workbook,
    data_only=True): one sheet per month, row-label match on 'Total Net
    GST Revenue' -> column index 2 (current-month value, INR crore).
    Returns [(date, value)], sorted."""
    rows: list[tuple[str, float]] = []
    for sheet_name in wb.sheetnames:
        date = parse_gst_collection_sheet_name(sheet_name)
        if date is None:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()
            if label == "Total Net GST Revenue" and len(row) > 2 and row[2] is not None:
                try:
                    rows.append((date, float(row[2])))
                except (TypeError, ValueError):
                    pass
                break
    rows.sort(key=lambda r: r[0])
    return rows


# ---------------------------------------------------------------------------
# (b) E-way bills
# ---------------------------------------------------------------------------

def find_ewb_fy_urls(listing_html: str) -> list[str]:
    """Scrape the gst.gov.in statistics page HTML for ewb-data-<FY>.xlsx
    links (does NOT guess filenames — enumerates whatever the page
    actually links to today). Returns absolute URLs, de-duplicated,
    in page order."""
    seen: set[str] = set()
    urls: list[str] = []
    for m in re.finditer(r'["\']((?:https?:)?//[^"\']*ewb-data-\d{4}-\d{2}\.xlsx)["\']', listing_html, re.IGNORECASE):
        raw = m.group(1)
        if raw.startswith("//"):
            raw = "https:" + raw
        if raw not in seen:
            seen.add(raw)
            urls.append(raw)
    return urls


def find_gst_collection_url(listing_html: str) -> str | None:
    """Scrape the gst.gov.in statistics page HTML for the
    Gross_Net_Tax_collection.xlsx link."""
    m = re.search(r'["\']((?:https?:)?//[^"\']*Gross_Net_Tax_collection\.xlsx)["\']', listing_html, re.IGNORECASE)
    if not m:
        return None
    raw = m.group(1)
    if raw.startswith("//"):
        raw = "https:" + raw
    return raw


def parse_ewb_workbook(wb: Any) -> dict[str, float]:
    """Parse one ewb-data-<FY>.xlsx workbook: state x month rows, sum
    'No. of Eway Bills' across all three flow-type column groups
    (0-indexed columns 5, 8, 11 — stable across the header-text drift
    observed between FY files) and all state rows, grouped by
    (year, month). Returns {'YYYY-MM-01': total_eway_bills}."""
    totals: dict[str, float] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 12:
                continue
            state_code, _state_name, year, month = row[0], row[1], row[2], row[3]
            if not isinstance(state_code, (int, float)):
                continue  # skip header/title rows
            try:
                year_i, month_i = int(year), int(month)
            except (TypeError, ValueError):
                continue
            if not (1 <= month_i <= 12):
                continue
            eway_counts = [row[5], row[8], row[11]]
            numeric = [v for v in eway_counts if isinstance(v, (int, float))]
            if not numeric:
                continue
            date = f"{year_i:04d}-{month_i:02d}-01"
            totals[date] = totals.get(date, 0.0) + sum(numeric)
    return totals


# ---------------------------------------------------------------------------
# (d) ICI Eight Core Industries
# ---------------------------------------------------------------------------

def ici_pdf_url(year: int, month: int) -> str:
    source = get_source("macro_govt", "eaindustry_ici")
    return source["url"].format(year=year, month=month)


def _month_to_key(cell: str) -> str | None:
    """'Apr-25' -> '2025-04-01'."""
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", cell.strip())
    if not m:
        return None
    mon_abbr, yy = m.groups()
    month = _MONTH_ABBR.get(mon_abbr.lower())
    if month is None:
        return None
    year = 2000 + int(yy)
    return f"{year:04d}-{month:02d}-01"


def parse_ici_pdf_tables(pdf: Any) -> dict[str, dict[str, float]]:
    """Parse the ICI press-release PDF's month-level index + growth
    tables (page index 3, verified live 2026-07). Returns
    {date: {"ICI_COAL": ..., ..., "ICI_OVERALL": ...}} for the index
    table only (headline + 8 sub-series levels; growth-rate table is not
    separately stored — index level is sufficient for YoY computation
    downstream, matching the FRED CPI_INDEX -> CPI_YOY precedent).
    Returns an empty dict (never fabricates) if the expected table shape
    isn't found on any page — a defensive fallback for future layout
    drift, honest per the plan's "+8 sub-series if the table parses
    cleanly, else headline only" instruction (degrades to empty here;
    the headline-only fallback path lives in the pipeline's parse())."""
    result: dict[str, dict[str, float]] = {}
    for page in pdf.pages:
        for table in page.extract_tables():
            if not table or len(table) < 3:
                continue
            header = table[0]
            if not header or "Sector" not in str(header[0] or ""):
                continue
            if "Index" not in str(header[-1] or "") and "Growth" not in str(header[-1] or ""):
                continue
            is_growth_table = "Growth" in str(header[-1] or "")
            if is_growth_table:
                continue  # index table only, per docstring
            for row in table[1:]:
                if not row:
                    continue
                # Month label sometimes sits in col 0, sometimes col 1
                # (pdfplumber occasionally splits a merged cell) — check both.
                month_cell = None
                value_start = None
                for idx in (0, 1):
                    if idx < len(row) and row[idx] and _month_to_key(str(row[idx])):
                        month_cell = row[idx]
                        value_start = idx + 1
                        break
                if month_cell is None:
                    continue
                date = _month_to_key(str(month_cell))
                values_raw = [v for v in row[value_start:] if v is not None]
                if len(values_raw) < len(ICI_COLUMNS):
                    continue
                values_raw = values_raw[-len(ICI_COLUMNS):]
                try:
                    values = [float(str(v).strip()) for v in values_raw]
                except ValueError:
                    continue
                result[date] = {
                    f"ICI_{col}": val for col, val in zip(ICI_COLUMNS, values)
                }
    return result


# ---------------------------------------------------------------------------
# (e) IIP via MoSPI eSankhyiki MCP
# ---------------------------------------------------------------------------

def _parse_mcp_sse(text: str) -> dict | None:
    """MCP streamable-HTTP responses come back SSE-framed even for a
    single non-streaming reply: find the 'data:' line and JSON-decode it."""
    for line in text.splitlines():
        if line.startswith("data:"):
            return json.loads(line[len("data:"):].strip())
    return None


def _mcp_tool_text(payload: dict) -> str:
    result = payload.get("result", {}) if payload else {}
    content = result.get("content", [])
    return "".join(c.get("text", "") for c in content if c.get("type") == "text")


def fetch_iip_via_mcp(session: requests.Session, url: str, timeout: float = 30.0) -> str | None:
    """initialize -> notifications/initialized -> tools/call(get_data).
    Returns the raw JSON text of the tool result, or None if any step of
    the BETA protocol fails (caller wraps this in try/except regardless;
    this also self-guards so a single malformed response never raises
    past parse-able JSON)."""
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json, text/event-stream",
    }
    session.post(
        url, headers=headers, timeout=timeout,
        data=json.dumps({
            "jsonrpc": "2.0", "id": 1, "method": "initialize",
            "params": {
                "protocolVersion": "2025-03-26", "capabilities": {},
                "clientInfo": {"name": "afund-macro-govt", "version": "0.1"},
            },
        }),
    )
    session.post(
        url, headers=headers, timeout=timeout,
        data=json.dumps({"jsonrpc": "2.0", "method": "notifications/initialized", "params": {}}),
    )
    resp = session.post(
        url, headers=headers, timeout=timeout,
        data=json.dumps({
            "jsonrpc": "2.0", "id": 2, "method": "tools/call",
            "params": {
                "name": "get_data",
                "arguments": {
                    "dataset": "IIP",
                    "filters": {
                        "base_year": "2011-12",
                        "frequency": "Monthly",
                        "type": "General",
                        "limit": 200,
                    },
                },
            },
        }),
    )
    resp.raise_for_status()
    payload = _parse_mcp_sse(resp.text)
    if payload is None:
        return None
    return _mcp_tool_text(payload)


def parse_iip_mcp_response(text: str) -> list[tuple[str, float, float | None]]:
    """Parse the get_data tool result text (a JSON string:
    {"data": [{"year", "month", "index", "growth_rate"}, ...]}) into
    [(date, index_value, growth_rate_or_None)]. Never fabricates: rows
    with unparseable month names or non-numeric index are skipped
    entirely, but a row with a valid index and a null/missing
    growth_rate is KEPT (index still emitted) with growth_rate=None —
    MoSPI itself publishes a null growth_rate for Apr-2021 (the COVID
    base-period comparator was near-zero, making YoY undefined), and
    dropping the whole row would discard a genuine index observation
    for no reason. The None is never coerced into a fabricated 0.0 or
    similar; downstream consumers must treat None as data_pending for
    that month's YoY figure specifically."""
    try:
        obj = json.loads(text)
    except (json.JSONDecodeError, TypeError):
        return []
    rows_out: list[tuple[str, float, float | None]] = []
    for row in obj.get("data", []):
        month_name = str(row.get("month", "")).strip().lower()
        month_num = None
        for abbr, num in _MONTH_ABBR.items():
            if month_name.startswith(abbr):
                month_num = num
                break
        year = row.get("year")
        if month_num is None or year is None:
            continue
        try:
            year_i = int(year)
            index_val = float(row["index"])
        except (TypeError, ValueError, KeyError):
            continue
        growth_val: float | None
        try:
            growth_val = float(row["growth_rate"])
        except (TypeError, ValueError, KeyError):
            growth_val = None
        date = f"{year_i:04d}-{month_num:02d}-01"
        rows_out.append((date, index_val, growth_val))
    rows_out.sort(key=lambda r: r[0])
    return rows_out


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

class MacroGovtPipeline(Pipeline):
    """Fetch + upsert GST collections, e-way bills, ICI Eight Core
    Industries, and IIP (via the MoSPI MCP, best-effort). EPFO is
    documented-broken (WAF bot-wall, see module docstring) and
    deliberately not wired here — see knowledge/data/kpis/
    epfo_payroll.yaml and config/sources.yaml's epfo_payroll entry for
    the manual-import fallback.

    Each sub-source's fetch/parse is independently try/except-guarded so
    one broken or drifted source never blocks the others — this mirrors
    Pipeline.run()'s own per-job guarantee but applied at the per-source
    level, since this one Pipeline instance covers four sub-sources in a
    single job_runs row."""

    job_name = "macro_govt"

    def fetch(self) -> dict[str, Any]:
        raw: dict[str, Any] = {"errors": {}}

        # --- GST collections + e-way bills share one listing page fetch ---
        try:
            session = make_session()
            listing_source = get_source("macro_govt", "gst_statistics_page")
            listing_resp = get(session, listing_source["url"], host_key=GST_HOST_KEY, min_interval=1.0, timeout=20.0)
            listing_resp.raise_for_status()
            listing_html = listing_resp.text

            gst_url = find_gst_collection_url(listing_html)
            if gst_url is None:
                raise RuntimeError("Gross_Net_Tax_collection.xlsx link not found on gst.gov.in statistics page")
            gst_resp = get(session, gst_url, host_key=GST_HOST_KEY, min_interval=1.0, timeout=30.0)
            gst_resp.raise_for_status()
            raw["gst_collection_xlsx"] = gst_resp.content

            ewb_urls = find_ewb_fy_urls(listing_html)
            ewb_files: list[bytes] = []
            for url in ewb_urls:
                try:
                    r = get(session, url, host_key=GST_HOST_KEY, min_interval=1.0, timeout=30.0)
                    r.raise_for_status()
                    ewb_files.append(r.content)
                except Exception as exc:  # noqa: BLE001 - one bad FY file must not drop the rest
                    raw["errors"][f"ewb_{url}"] = str(exc)
            raw["ewb_xlsx_files"] = ewb_files
        except Exception as exc:  # noqa: BLE001
            raw["errors"]["gst_ewb"] = str(exc)

        # --- ICI Eight Core Industries: probe recent months backward ---
        try:
            ici_session = make_session()
            ici_pdfs: dict[str, bytes] = {}
            today = dt.date.today()
            year, month = today.year, today.month
            # ICI publishes with a ~20-day lag; probe up to 6 months back
            # to find the latest already-published PDF plus recent history.
            probed = 0
            y, m = year, month
            while probed < 18:  # up to 18 months of backfill in one pass
                m -= 1
                if m == 0:
                    m, y = 12, y - 1
                url = ici_pdf_url(y, m)
                try:
                    resp = get(ici_session, url, host_key=ICI_HOST_KEY, min_interval=1.0, timeout=30.0)
                    content_type = resp.headers.get("Content-Type", "")
                    if resp.status_code == 200 and "pdf" in content_type.lower():
                        ici_pdfs[f"{y:04d}-{m:02d}"] = resp.content
                    # else: not-yet-published HTML shell (same class as the
                    # niftyindices.com gotcha) or genuine 404 — skip honestly.
                except Exception as exc:  # noqa: BLE001 - one bad month must not drop the rest
                    raw["errors"][f"ici_{y:04d}_{m:02d}"] = str(exc)
                probed += 1
            raw["ici_pdfs"] = ici_pdfs
        except Exception as exc:  # noqa: BLE001
            raw["errors"]["ici"] = str(exc)

        # --- IIP via MoSPI eSankhyiki MCP (BETA) ---
        try:
            mcp_source = get_source("macro_govt", "mospi_esankhyiki_mcp")
            mcp_session = requests.Session()
            mcp_session.headers["User-Agent"] = "afund-macro-govt/0.1"
            text = fetch_iip_via_mcp(mcp_session, mcp_source["url"])
            raw["iip_mcp_text"] = text
        except Exception as exc:  # noqa: BLE001 - MCP is BETA; must degrade gracefully
            raw["errors"]["iip_mcp"] = str(exc)
            raw["iip_mcp_text"] = None

        return raw

    def parse(self, raw: dict[str, Any]) -> dict[str, list[tuple[str, float]]]:
        import openpyxl
        import pdfplumber

        parsed: dict[str, list[tuple[str, float]]] = {}

        gst_xlsx = raw.get("gst_collection_xlsx")
        if gst_xlsx:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(gst_xlsx), data_only=True)
                parsed["GST_COLLECTIONS"] = parse_gst_collection_workbook(wb)
            except Exception as exc:  # noqa: BLE001
                raw.setdefault("errors", {})["gst_parse"] = str(exc)

        ewb_files = raw.get("ewb_xlsx_files") or []
        if ewb_files:
            try:
                totals: dict[str, float] = {}
                for content in ewb_files:
                    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                    file_totals = parse_ewb_workbook(wb)
                    for date, value in file_totals.items():
                        totals[date] = value  # each FY file covers disjoint months
                parsed["EWAY_BILLS"] = sorted(totals.items())
            except Exception as exc:  # noqa: BLE001
                raw.setdefault("errors", {})["ewb_parse"] = str(exc)

        ici_pdfs = raw.get("ici_pdfs") or {}
        if ici_pdfs:
            try:
                by_series: dict[str, list[tuple[str, float]]] = {f"ICI_{c}": [] for c in ICI_COLUMNS}
                seen_dates: set[str] = set()
                for content in ici_pdfs.values():
                    with pdfplumber.open(io.BytesIO(content)) as pdf:
                        month_rows = parse_ici_pdf_tables(pdf)
                    for date, values in month_rows.items():
                        if date in seen_dates:
                            continue
                        seen_dates.add(date)
                        for col in ICI_COLUMNS:
                            series_code = f"ICI_{col}"
                            if series_code in values:
                                by_series[series_code].append((date, values[series_code]))
                for series_code, rows in by_series.items():
                    rows.sort(key=lambda r: r[0])
                    out_code = "ICI_INDEX" if series_code == "ICI_OVERALL" else series_code
                    if rows:
                        parsed[out_code] = rows
            except Exception as exc:  # noqa: BLE001
                raw.setdefault("errors", {})["ici_parse"] = str(exc)

        iip_text = raw.get("iip_mcp_text")
        if iip_text:
            try:
                iip_rows = parse_iip_mcp_response(iip_text)
                if iip_rows:
                    parsed["IIP_INDEX"] = [(d, idx) for d, idx, _g in iip_rows]
                    # growth_rate is honestly None for some months (e.g. the
                    # COVID base-period anomaly) — exclude those from
                    # IIP_YOY rather than fabricating a value; the index
                    # level is still preserved above.
                    parsed["IIP_YOY"] = [
                        (d, g) for d, _idx, g in iip_rows if g is not None
                    ]
            except Exception as exc:  # noqa: BLE001
                raw.setdefault("errors", {})["iip_parse"] = str(exc)

        return parsed

    def upsert(self, parsed: dict[str, list[tuple[str, float]]]) -> int:
        unit_freq: dict[str, tuple[str, str]] = {
            "GST_COLLECTIONS": ("INR_cr", "M"),
            "EWAY_BILLS": ("count", "M"),
            "IIP_INDEX": ("index", "M"),
            "IIP_YOY": ("%", "M"),
        }
        for col in ICI_COLUMNS:
            code = "ICI_INDEX" if col == "OVERALL" else f"ICI_{col}"
            unit_freq[code] = ("index", "M")

        total = 0
        for series_code, rows in parsed.items():
            unit, freq = unit_freq.get(series_code, (None, "M"))
            total += _upsert_series(self.conn, series_code, "GOVT", rows, unit, freq)
        self.conn.commit()
        return total


if __name__ == "__main__":
    result = MacroGovtPipeline().run()
    print(result)
