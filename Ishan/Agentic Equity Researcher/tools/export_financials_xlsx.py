"""Excel export of a ticker's extracted financials — deterministic, zero
tokens (openpyxl). Reads workspace/<TICKER>/ directly and writes
workspace/<TICKER>/exports/<TICKER>_financials.xlsx.

Ported verbatim from the fund repo (D:\\Documents\\Claude\\1Projects\\AI
Native Fund, research/equity_researcher/tools/export_financials_xlsx.py) —
no fund imports there either, so no adaptation was needed for this
standalone copy.

Sheets:
  IS_tree / BS_tree / CF_tree   3-level indented line-item tree (mirrors
                                 state/comprehensive_statement.json), one
                                 sheet per statement, periods as columns,
                                 indentation encoded via cell indent level 1:1
                                 with the tree's `level` field.
  Quarterly                     same tree data but restricted to Q/H period
                                 columns only (the *_tree sheets show every
                                 period; this sheet is the quarter-only cut).
  Ratios                        computed ratios from facts/derived_metrics.json
                                 (metric x FY, mirrors render_tables.py's
                                 ratio_summary spec).
  EPS_Bridge                    mirrors state/eps_bridge_check.json — one row
                                 per rule_id with status/value/threshold/note.
  RedFlags                      mirrors state/red_flags.json ledger.

Usage:
  python tools/export_financials_xlsx.py workspace/TICKER
      [--out workspace/TICKER/exports/TICKER_financials.xlsx]

No fund imports (standalone ER must stay self-contained).
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)


def _load_json(path: Path, default):
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return default


# --- tree sheets (IS_tree / BS_tree / CF_tree / Quarterly) -----------------

def _all_periods(tree: dict) -> list[str]:
    periods: set[str] = set()

    def walk(node: dict):
        periods.update((node.get("values") or {}).keys())
        for c in node.get("children") or []:
            walk(c)

    for root in tree.values():
        walk(root)
    return sorted(periods, key=lambda p: (0 if p.startswith("FY") and len(p) == 6 else 1, p))


def _cell_safe(v):
    """Excel cells hold scalars. A structured ledger field (a why-chain list, a merge-discrepancy
    dict) used to abort the ENTIRE workbook with `ValueError: Cannot convert {...} to Excel` - one
    field shape taking out every sheet. Render it instead: this export is a reading surface, and a
    readable summary beats no workbook at all."""
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    if isinstance(v, (list, tuple)):
        return "; ".join(str(x) for x in v)
    if isinstance(v, dict):
        return "; ".join(f"{k}={v[k]}" for k in list(v)[:6])
    return str(v)


def _fmt_cell_value(entry: dict | None):
    if entry is None:
        return None
    v = entry.get("value")
    if isinstance(v, (int, float)):
        return v
    return v  # leave strings / None as-is; openpyxl handles both


def _write_tree_sheet(ws, tree: dict, periods: list[str], basis: str | None = None):
    ws.append(["Line item"] + periods)
    for c in ws[1]:
        c.font = HEADER_FONT

    def write_node(node: dict, row_cursor: list[int]):
        row_cursor[0] += 1
        r = row_cursor[0]
        level = node.get("level", 1) or 1
        label_cell = ws.cell(row=r, column=1, value=node.get("label") or node.get("metric"))
        label_cell.alignment = Alignment(indent=max(level - 1, 0) * 2)
        values = node.get("values") or {}
        for ci, period in enumerate(periods, start=2):
            entry = values.get(period)
            if entry is not None and basis is not None and entry.get("basis") not in (basis, "na"):
                entry = None
            ws.cell(row=r, column=ci, value=_fmt_cell_value(entry))
        for child in sorted(node.get("children") or [], key=_order_key):
            write_node(child, row_cursor)

    row_cursor = [1]
    for root in sorted(tree.values(), key=_order_key):
        write_node(root, row_cursor)

    ws.column_dimensions["A"].width = 42
    for ci in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _dominant_basis(tree: dict) -> str | None:
    """The basis that spans the most FISCAL YEARS - not the one with the most facts.

    Counting facts picked standalone for NALCO (157 vs 122) even though consolidated covers six
    fiscal years and standalone covers one, so the single-basis column filter blanked almost the
    whole statement. The filter exists to produce a COMPARABLE column series, so coverage across
    periods is the right measure. Ties fall to consolidated, the basis a reader expects.
    """
    fy_by_basis: dict[str, set] = {}

    def walk(node: dict):
        for period, entry in (node.get("values") or {}).items():
            b = entry.get("basis")
            if b and b != "na" and period.startswith("FY") and len(period) == 6:
                fy_by_basis.setdefault(b, set()).add(period)
        for c in node.get("children") or []:
            walk(c)

    for root in tree.values():
        walk(root)
    if not fy_by_basis:
        return None
    return max(fy_by_basis, key=lambda b: (len(fy_by_basis[b]), b == "consolidated"))


def _add_statement_sheet(wb: Workbook, sheet_name: str, tree: dict):
    ws = wb.create_sheet(sheet_name)
    if not tree:
        ws.append(["No data — comprehensive_statement.json has no entries for this statement."])
        return
    periods = _all_periods(tree)
    basis = _dominant_basis(tree)
    _write_tree_sheet(ws, tree, periods, basis=basis)


def _add_quarterly_sheet(wb: Workbook, comp_statement: dict):
    ws = wb.create_sheet("Quarterly")
    q_periods: set[str] = set()

    def collect_periods(tree: dict):
        def walk(node):
            for p in (node.get("values") or {}):
                if not (p.startswith("FY") and len(p) == 6):  # exclude plain FYxxxx annual columns
                    q_periods.add(p)
            for c in node.get("children") or []:
                walk(c)
        for root in tree.values():
            walk(root)

    for stmt_key in ("income_statement", "balance_sheet", "cash_flow"):
        collect_periods(comp_statement.get(stmt_key, {}) or {})

    periods = sorted(q_periods)
    if not periods:
        ws.append(["No quarterly/half-year periods found in comprehensive_statement.json."])
        return

    ws.append(["Statement", "Line item"] + periods)
    for c in ws[1]:
        c.font = HEADER_FONT

    row = 1
    for stmt_key, stmt_label in (
        ("income_statement", "Income Statement"),
        ("balance_sheet", "Balance Sheet"),
        ("cash_flow", "Cash Flow"),
    ):
        tree = comp_statement.get(stmt_key, {}) or {}
        basis = _dominant_basis(tree)

        def write_node(node: dict):
            nonlocal row
            row += 1
            values = node.get("values") or {}
            level = node.get("level", 1) or 1
            label_cell = ws.cell(row=row, column=2, value=node.get("label") or node.get("metric"))
            label_cell.alignment = Alignment(indent=max(level - 1, 0) * 2)
            ws.cell(row=row, column=1, value=stmt_label)
            for ci, period in enumerate(periods, start=3):
                entry = values.get(period)
                if entry is not None and basis is not None and entry.get("basis") not in (basis, "na"):
                    entry = None
                ws.cell(row=row, column=ci, value=_fmt_cell_value(entry))
            for child in sorted(node.get("children") or [], key=_order_key):
                write_node(child)

        for root in sorted(tree.values(), key=_order_key):
            write_node(root)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    for ci in range(3, len(periods) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# --- Ratios sheet ------------------------------------------------------------

RATIO_METRICS = [
    ("gross_margin", "Gross margin %"),
    ("ebitda_margin", "EBITDA margin %"),
    ("net_margin", "PAT margin %"),
    ("roe", "ROE %"),
    ("roce", "ROCE %"),
    ("asset_turnover", "Asset turnover (x)"),
    ("dso_days", "Receivable days"),
    ("inventory_days", "Inventory days"),
    ("payable_days", "Payable days"),
    ("ccc_days", "Cash conversion cycle (days)"),
    ("current_ratio", "Current ratio (x)"),
    ("debt_equity", "Debt/Equity (x)"),
    ("interest_coverage", "Interest coverage (x)"),
    ("net_debt_ebitda", "Net debt/EBITDA (x)"),
    ("cfo_to_ebitda", "CFO/EBITDA %"),
    ("fcf", "Free cash flow (INR cr)"),
]


def _add_ratios_sheet(wb: Workbook, derived_facts: list[dict]):
    ws = wb.create_sheet("Ratios")
    live = [f for f in derived_facts if "superseded" not in (f.get("flags") or [])]
    periods = sorted({f.get("period") for f in live if f.get("period") and f.get("period").startswith("FY") and len(f.get("period")) == 6})
    if not periods:
        ws.append(["No derived_metrics.json ratio facts found (run tools/compute_ratios.py first)."])
        return

    idx: dict[tuple, dict] = {}
    for f in live:
        key = (f.get("metric"), f.get("period"))
        idx.setdefault(key, f)  # first wins; derived facts have no reported/computed collision here

    ws.append(["Ratio"] + periods)
    for c in ws[1]:
        c.font = HEADER_FONT
    for metric, label in RATIO_METRICS:
        row = [label]
        for p in periods:
            f = idx.get((metric, p))
            row.append(f.get("value") if f else None)
        ws.append(row)

    ws.column_dimensions["A"].width = 32
    for ci in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14


# --- EPS_Bridge sheet ---------------------------------------------------------

def _add_eps_bridge_sheet(wb: Workbook, eps_bridge_check: dict):
    ws = wb.create_sheet("EPS_Bridge")
    ws.append(["Rule", "Status", "Value", "Threshold", "Note"])
    for c in ws[1]:
        c.font = HEADER_FONT
    if not eps_bridge_check:
        ws.append(["No eps_bridge_check.json found — run tools/eps_bridge_check.py first.", "", "", "", ""])
        return
    for rule_id, result in eps_bridge_check.items():
        if rule_id.startswith("_"):
            continue  # skip _basis / _periods metadata keys
        if not isinstance(result, dict):
            continue
        value = result.get("value")
        value_str = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
        ws.append([rule_id, result.get("status"), value_str, result.get("threshold"), result.get("note")])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 60
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True)


# --- RedFlags sheet ------------------------------------------------------------

RED_FLAG_COLUMNS = ["id", "category", "flag", "status", "severity", "confidence", "threshold", "owner"]


def _add_red_flags_sheet(wb: Workbook, red_flags: list[dict]):
    ws = wb.create_sheet("RedFlags")
    ws.append([c.replace("_", " ").title() for c in RED_FLAG_COLUMNS])
    for c in ws[1]:
        c.font = HEADER_FONT
    if not red_flags:
        ws.append(["No red_flags.json entries found." if True else ""] + [""] * (len(RED_FLAG_COLUMNS) - 1))
        return
    for entry in red_flags:
        ws.append([_cell_safe(entry.get(col)) for col in RED_FLAG_COLUMNS])

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["C"].width = 60
    for ci in range(1, len(RED_FLAG_COLUMNS) + 1):
        letter = get_column_letter(ci)
        if letter not in ("A", "C"):
            ws.column_dimensions[letter].width = 16


# --- orchestration -------------------------------------------------------------

# --- horizontal / vertical analysis ---------------------------------------
# The numbers are needed in the analysis, but the Excel must carry them too. Both are computed
# HERE from the same tree the *_tree sheets render, so any percentage ties back to a line item and
# period without leaving the workbook.
#
#   Horizontal (trend) analysis = YoY % change per line item, period over period.
#   Vertical (common-size)      = each line as a % of its statement base:
#                                   income statement -> revenue from operations
#                                   balance sheet    -> total assets
#                                   cash flow        -> cash from operating activities
# The base is resolved by metric-name match and NAMED in the sheet header. If it cannot be found
# the sheet says so rather than dividing by an arbitrary denominator.

_VERTICAL_BASE = {
    "IS": (("revenue_from_operations", "total_income", "revenue"), "revenue from operations"),
    "BS": (("total_assets", "total_equity_and_liabilities"), "total assets"),
    "CF": (("cash_from_operating_activities", "cfo"), "cash from operating activities"),
}


# --- statement ordering ---------------------------------------------------
# Sheets used to sort by label, which opened the income statement on "Current tax assets (Net)".
# A financial statement is read top-to-bottom, so rows follow the canonical sequence below;
# unrecognised metrics keep alphabetical order AFTER the known ones, so nothing is dropped and no
# position is invented for a line we cannot place.
_STATEMENT_ORDER = (
    # income statement
    "revenue_from_operations", "other_income", "total_income",
    "cost_of_raw_materials", "cost_of_materials", "purchases_stock_in_trade",
    "changes_in_inventories", "employee_benefits", "power_and_fuel", "power_fuel",
    "freight", "sub_contracting", "other_expenses", "total_expenses",
    "ebitda", "depreciation", "amortization", "amortisation", "finance_costs",
    "profit_before_exceptional", "exceptional_item", "profit_before_tax", "pbt",
    "current_tax", "deferred_tax", "total_tax_expense",
    "profit_for_the_year", "pat", "share_of_profit", "minority_interest",
    "other_comprehensive_income", "total_comprehensive_income", "tci",
    "eps", "weighted_shares",
    # balance sheet - assets then equity then liabilities
    "property_plant", "ppe", "cwip", "capital_work", "right_of_use",
    "goodwill", "intangible", "investments", "loans_noncurrent",
    "other_financial_assets_noncurrent", "other_noncurrent",
    "inventories", "trade_receivables", "cash_and_cash_equivalents",
    "cash_and_equivalents", "cash_and_bank", "bank_balances", "loans_current",
    "other_financial_assets_current", "other_current_assets",
    "current_assets", "total_assets",
    "equity_share_capital", "share_capital", "other_equity", "reserves", "total_equity",
    "borrowings", "lease_liabilit", "trade_payables",
    "other_financial_liabilities", "provisions", "current_liabilities",
    "total_liabilities", "contingent_liabilities", "net_debt", "net_worth",
    # cash flow
    "cash_from_operating", "cfo", "working_capital_movement",
    "cash_from_investing", "cfi", "purchase_of_ppe", "purchase_of_intangibles",
    "sale_of_ppe", "net_capex", "capex", "cash_from_financing", "cff",
    "proceeds_from", "repayment_of", "interest_paid", "dividend_paid", "fcf",
)


def _order_key(node: dict):
    """(rank, label) - known face lines in statement order, then the rest alphabetically."""
    metric = (node.get("metric") or "").lower()
    label = node.get("label") or metric
    for i, key in enumerate(_STATEMENT_ORDER):
        if metric.startswith(key) or key in metric:
            return (i, label)
    # Bucket nodes always last: they hold items whose position is unknown by definition.
    if metric.startswith("_unattributed_level"):
        return (len(_STATEMENT_ORDER) + 1, label)
    return (len(_STATEMENT_ORDER), label)


def _flatten(tree: dict) -> list[dict]:
    """Depth-first node list, preserving the tree display order and level."""
    out: list[dict] = []

    def walk(node: dict):
        out.append(node)
        for child in sorted(node.get("children") or [], key=_order_key):
            walk(child)

    for root in sorted(tree.values(), key=_order_key):
        walk(root)
    return out


def _fy_only(periods: list[str]) -> list[str]:
    """Fiscal years only. YoY or common-size across a mix of FY and quarterly columns would
    compare unlike periods."""
    return [q for q in periods if q.startswith("FY") and len(q) == 6]


def _num_at(node: dict, period: str, basis: str | None):
    entry = (node.get("values") or {}).get(period)
    if entry is None:
        return None
    if basis is not None and entry.get("basis") not in (basis, "na"):
        return None
    v = entry.get("value")
    return v if isinstance(v, (int, float)) else None


def _off_basis_count(tree: dict, basis: str | None) -> int:
    """Line items that have FY data, but none of it on the chosen basis - so every one of their
    cells is blank for a reason the sheet should state rather than leave to inference."""
    if not basis:
        return 0
    n = 0
    for node in _flatten(tree):
        fy = {p: e for p, e in (node.get("values") or {}).items()
              if p.startswith("FY") and len(p) == 6}
        if fy and not any(e.get("basis") in (basis, "na") for e in fy.values()):
            n += 1
    return n


def _add_horizontal_sheet(wb: Workbook, sheet_name: str, tree: dict):
    ws = wb.create_sheet(sheet_name)
    if not tree:
        ws.append(["No data - comprehensive_statement.json has no entries for this statement."])
        return
    periods = _fy_only(_all_periods(tree))
    basis = _dominant_basis(tree)
    if len(periods) < 2:
        ws.append([f"Needs at least two fiscal years to compute YoY; found {len(periods)}."])
        return
    off = _off_basis_count(tree, basis)
    ws.append([f"Horizontal (trend) analysis - YoY % change. Basis: {basis or 'mixed'} "
               f"(chosen as the basis covering the most fiscal years). Blank means the line item is "
               f"missing in one of the two periods"
               + (f", or reported only on the other basis - {off} line item(s) here have FY data but "
                  f"none of it on the {basis} basis, and are deliberately left blank rather than "
                  f"blended across bases." if off else ".")])
    ws["A1"].font = TITLE_FONT
    ws.append(["Line item"] + [f"{periods[i]} vs {periods[i-1]}" for i in range(1, len(periods))])
    for c in ws[2]:
        c.font = HEADER_FONT

    for node in _flatten(tree):
        row = [node.get("label") or node.get("metric")]
        for i in range(1, len(periods)):
            prev = _num_at(node, periods[i - 1], basis)
            cur = _num_at(node, periods[i], basis)
            row.append(round((cur / prev - 1) * 100, 1)
                       if (prev not in (None, 0) and cur is not None) else None)
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(
            indent=max((node.get("level", 1) or 1) - 1, 0) * 2)

    ws.column_dimensions["A"].width = 42
    for ci in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "B3"


def _add_vertical_sheet(wb: Workbook, sheet_name: str, tree: dict, which: str):
    ws = wb.create_sheet(sheet_name)
    if not tree:
        ws.append(["No data - comprehensive_statement.json has no entries for this statement."])
        return
    periods = _fy_only(_all_periods(tree))
    basis = _dominant_basis(tree)
    candidates, base_label = _VERTICAL_BASE[which]

    flat = _flatten(tree)
    base_node = None
    for cand in candidates:                     # first candidate that actually resolves
        for node in flat:
            if node.get("metric") == cand:
                base_node = node
                break
        if base_node is not None:
            base_label = cand
            break
    if base_node is None:
        ws.append([f"Common-size base not found. Looked for {list(candidates)}; none is present in "
                   f"this statement, so every line would divide by an arbitrary denominator. "
                   f"Extract the base line and re-run rather than reading a guessed percentage."])
        return

    off = _off_basis_count(tree, basis)
    ws.append([f"Vertical (common-size) analysis - each line as % of {base_label}. "
               f"Basis: {basis or 'mixed'} (chosen as the basis covering the most fiscal years)."
               + (f" {off} line item(s) have FY data only on the other basis and are left blank: "
                  f"blending bases in one common-size column would be arithmetically wrong."
                  if off else "")])
    ws["A1"].font = TITLE_FONT
    ws.append(["Line item"] + periods)
    for c in ws[2]:
        c.font = HEADER_FONT

    for node in flat:
        row = [node.get("label") or node.get("metric")]
        for period in periods:
            base = _num_at(base_node, period, basis)
            val = _num_at(node, period, basis)
            row.append(round(val / base * 100, 1)
                       if (base not in (None, 0) and val is not None) else None)
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(
            indent=max((node.get("level", 1) or 1) - 1, 0) * 2)

    ws.column_dimensions["A"].width = 42
    for ci in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = "B3"


def _count_inferred(comp_statement: dict) -> int:
    n = 0

    def walk(node: dict):
        nonlocal n
        if node.get("inferred_parent"):
            n += 1
        for c in node.get("children") or []:
            walk(c)

    for stmt in comp_statement.values():
        if isinstance(stmt, dict):
            for root in stmt.values():
                if isinstance(root, dict):
                    walk(root)
    return n


def _add_contents_sheet(wb: Workbook, ticker: str, notes: list[str]):
    """First tab: what every other tab holds, so the workbook navigates without this source file."""
    ws = wb.create_sheet("Contents", 0)
    ws.append([f"{ticker} - extracted financials"])
    ws["A1"].font = TITLE_FONT
    ws.append([])
    ws.append(["Tab", "What it holds"])
    for c in ws[3]:
        c.font = HEADER_FONT
    for tab, desc in [
        ("IS_tree", "Income statement, 3-level indented line-item tree x all periods"),
        ("BS_tree", "Balance sheet, same structure"),
        ("CF_tree", "Cash flow, same structure"),
        ("IS_horizontal", "Income statement YoY % change per line item (trend analysis)"),
        ("BS_horizontal", "Balance sheet YoY % change per line item"),
        ("CF_horizontal", "Cash flow YoY % change per line item"),
        ("IS_vertical", "Income statement common-size: each line as % of revenue"),
        ("BS_vertical", "Balance sheet common-size: each line as % of total assets"),
        ("CF_vertical", "Cash flow common-size: each line as % of cash from operations"),
        ("Other_metrics", "Production, capacity, dividends, buybacks - outside the three statements"),
        ("Quarterly", "The same tree restricted to quarterly / half-year columns"),
        ("Ratios", "Computed ratios from facts/derived_metrics.json"),
        ("EPS_Bridge", "Buy-side EPS-bridge rule verdicts from state/eps_bridge_check.json"),
        ("RedFlags", "The adjudicated red-flag ledger from state/red_flags.json"),
    ]:
        ws.append([tab, desc])
    if notes:
        ws.append([])
        ws.append(["Read this before using the numbers"])
        ws.cell(row=ws.max_row, column=1).font = HEADER_FONT
        for n in notes:
            ws.append([n])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96


def export_ticker(workspace_dir: Path, out_path: Path | None = None) -> Path:
    workspace_dir = Path(workspace_dir)
    ticker = workspace_dir.name

    comp_statement = _load_json(workspace_dir / "state" / "comprehensive_statement.json", {})
    derived = _load_json(workspace_dir / "facts" / "derived_metrics.json", {})
    derived_facts = derived.get("facts", derived) if isinstance(derived, dict) else (derived or [])
    eps_bridge_check = _load_json(workspace_dir / "state" / "eps_bridge_check.json", {})
    red_flags = _load_json(workspace_dir / "state" / "red_flags.json", [])
    if isinstance(red_flags, dict):
        red_flags = red_flags.get("flags", red_flags.get("red_flags", []))

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet; we add named ones below

    is_tree = comp_statement.get("income_statement", {}) or {}
    bs_tree = comp_statement.get("balance_sheet", {}) or {}
    cf_tree = comp_statement.get("cash_flow", {}) or {}
    other_tree = comp_statement.get("unclassified", {}) or {}

    _add_statement_sheet(wb, "IS_tree", is_tree)
    _add_statement_sheet(wb, "BS_tree", bs_tree)
    _add_statement_sheet(wb, "CF_tree", cf_tree)

    # Horizontal (YoY) and vertical (common-size) analysis per statement. Previously absent from
    # the workbook entirely: the Ratios sheet carried only a fixed ratio list, so a reader could not
    # see the statement-wide trend or mix that every financial-analysis section rests on.
    _add_horizontal_sheet(wb, "IS_horizontal", is_tree)
    _add_horizontal_sheet(wb, "BS_horizontal", bs_tree)
    _add_horizontal_sheet(wb, "CF_horizontal", cf_tree)
    _add_vertical_sheet(wb, "IS_vertical", is_tree, "IS")
    _add_vertical_sheet(wb, "BS_vertical", bs_tree, "BS")
    _add_vertical_sheet(wb, "CF_vertical", cf_tree, "CF")

    # Production / capacity / dividend metrics sit outside the three statements but are exactly what
    # the operating-KPI layer needs, so they get a tab rather than being dropped from the export.
    _add_statement_sheet(wb, "Other_metrics", other_tree)

    _add_quarterly_sheet(wb, comp_statement)
    _add_ratios_sheet(wb, derived_facts)
    _add_eps_bridge_sheet(wb, eps_bridge_check)
    _add_red_flags_sheet(wb, red_flags)

    notes: list[str] = []
    inferred = _count_inferred(comp_statement)
    if inferred:
        notes.append(f"{inferred} line item(s) have an INFERRED parent - extraction did not record "
                     f"`parent`, so the builder reconstructed the edge from the metric-name "
                     f"hierarchy. Indentation on those rows is a reconstruction, not a disclosure.")
    bucketed = sum(len(n.get("children") or []) for stmt in comp_statement.values()
                   if isinstance(stmt, dict)
                   for k, n in stmt.items() if str(k).startswith("_unattributed_level"))
    if bucketed:
        notes.append(f"{bucketed} line item(s) sit under a 'Level-N items whose parent line was not "
                     f"captured' node: real numbers whose position in the statement is unknown.")
    notes.append("Horizontal and vertical sheets use fiscal years only - mixing FY with quarterly "
                 "columns would compare unlike periods.")
    _add_contents_sheet(wb, ticker, notes)

    if out_path is None:
        out_path = workspace_dir / "exports" / f"{ticker}_financials.xlsx"
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workspace_dir", help="workspace/TICKER directory")
    ap.add_argument("--out", default=None, help="output xlsx path (default: workspace/TICKER/exports/TICKER_financials.xlsx)")
    a = ap.parse_args()

    out_path = export_ticker(Path(a.workspace_dir), Path(a.out) if a.out else None)
    print(f"OK: financials export -> {out_path}")


if __name__ == "__main__":
    main()
