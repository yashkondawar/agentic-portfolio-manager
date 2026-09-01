"""Tests for the qtr_results Excel dossier: tax ledger, metrics and workbook."""

from __future__ import annotations

from datetime import date
from types import SimpleNamespace

import pandas as pd
import pytest

from backtesting.gfs.taxes import TaxConfig
from backtesting.qtr_results import dossier as D
from backtesting.qtr_results.config import live_mirror_config
from backtesting.qtr_results.portfolio import Portfolio, Position
from qtr_results import config as live_config


def _trade(symbol, entry, exit_, pnl, held):
    return SimpleNamespace(
        symbol=symbol, entry_date=entry, exit_date=exit_, pnl=pnl,
        holding_days=held, quantity=10, entry_price=100.0, exit_price=110.0,
        pnl_pct=10.0, exit_reason="target", sector="Tech", gross_pnl=pnl,
        costs=5.0,
    )


# ── Config preset ────────────────────────────────────────────────────────────


def test_live_mirror_tracks_live_config_not_backtest_defaults():
    cfg = live_mirror_config()
    assert cfg.risk_per_trade_pct == live_config.RISK_PER_TRADE_PCT == 4.0
    assert cfg.static_target_tiers == tuple(live_config.STATIC_TARGET_TIERS)
    assert cfg.static_target_tiers[0][1] == 20.0  # not the backtest's halved 10%
    assert cfg.max_holding_days == live_config.MAX_HOLDING_DAYS
    assert cfg.commission_pct == live_config.COMMISSION_PCT
    assert cfg.disable_profit_target is True
    # Research-only switches must stay off or the dossier stops describing live.
    assert not cfg.regime_filter and not cfg.use_sue and not cfg.anticipation_mode


def test_live_mirror_rejects_unknown_override():
    with pytest.raises(AttributeError):
        live_mirror_config(not_a_real_field=1)


def test_live_mirror_applies_overrides():
    cfg = live_mirror_config(starting_capital=1_000_000.0)
    assert cfg.starting_capital == 1_000_000.0


# ── Tax ledger ───────────────────────────────────────────────────────────────


def test_short_term_loss_carries_forward_and_shelters_next_year():
    cfg = TaxConfig()
    trades = [
        _trade("A", date(2023, 5, 1), date(2023, 9, 1), -50_000.0, 123),
        _trade("B", date(2023, 6, 1), date(2023, 10, 1), 10_000.0, 122),
        _trade("C", date(2024, 5, 1), date(2024, 9, 1), 200_000.0, 123),
    ]
    ledger = D.build_tax_ledger(D.classify_trades(trades, cfg), cfg)
    first, second = ledger.iloc[0], ledger.iloc[1]

    assert first["financial year"] == "FY2023-24"
    assert first["loss carried forward"] == pytest.approx(40_000.0)
    assert first["total tax paid"] == 0.0

    assert second["loss brought forward"] == pytest.approx(40_000.0)
    assert second["taxable ST"] == pytest.approx(160_000.0)
    # Sold after 23 Jul 2024, so the 20% rate applies, not the legacy 15%.
    assert second["tax on ST"] == pytest.approx(32_000.0)


def test_long_term_gain_gets_the_annual_exemption():
    cfg = TaxConfig()
    trades = [_trade("D", date(2023, 1, 1), date(2024, 6, 1), 300_000.0, 517)]
    ledger = D.build_tax_ledger(D.classify_trades(trades, cfg), cfg)
    row = ledger.iloc[0]
    assert row["long-term gain"] == pytest.approx(300_000.0)
    assert row["taxable LT"] == pytest.approx(300_000.0 - cfg.ltcg_exempt_per_year)
    assert row["tax on LT"] == pytest.approx(175_000.0 * 0.125)


def test_stcg_rate_changeover_is_respected():
    cfg = TaxConfig()
    before = D.build_tax_ledger(
        D.classify_trades([_trade("E", date(2024, 4, 1), date(2024, 7, 1), 100_000.0, 91)], cfg),
        cfg,
    )
    assert before.iloc[0]["tax on ST"] == pytest.approx(15_000.0)  # legacy 15%


def test_long_term_loss_cannot_shelter_short_term_gain():
    cfg = TaxConfig()
    trades = [
        _trade("F", date(2022, 1, 1), date(2023, 5, 1), -100_000.0, 485),
        _trade("G", date(2023, 6, 1), date(2023, 8, 1), 50_000.0, 61),
    ]
    ledger = D.build_tax_ledger(D.classify_trades(trades, cfg), cfg)
    row = ledger.iloc[0]
    assert row["taxable ST"] == pytest.approx(50_000.0)
    assert row["loss carried forward"] == pytest.approx(100_000.0)


def test_empty_ledger_keeps_reference_columns():
    ledger = D.build_tax_ledger([], TaxConfig())
    assert ledger.empty
    assert "total tax paid" in ledger.columns


def test_fy_end_maps_to_march_31():
    assert D.fy_end("FY2024-25") == date(2025, 3, 31)


# ── Metrics ──────────────────────────────────────────────────────────────────


def test_cagr_and_drawdown():
    assert D.cagr([100.0, 200.0], date(2020, 1, 1), date(2022, 1, 1)) == pytest.approx(41.39, abs=0.05)
    assert D.max_drawdown([100, 120, 60, 90]) == pytest.approx(-50.0)
    assert D.drawdown_series([100, 120, 60])[-1] == pytest.approx(-50.0)


def test_beta_of_a_series_against_itself_is_one():
    import numpy as np
    rets = np.linspace(-0.02, 0.03, 120)
    stats = D.beta_alpha(rets, rets, 10.0, 10.0)
    assert stats["beta"] == pytest.approx(1.0)
    assert stats["correlation"] == pytest.approx(1.0)
    assert stats["alpha"] == pytest.approx(0.0)


def test_metric_helpers_tolerate_degenerate_input():
    assert D.cagr([], date(2020, 1, 1), date(2021, 1, 1)) is None
    assert D.sharpe(D.daily_returns([100.0])) is None
    assert D.sortino(D.daily_returns([100.0, 101.0])) is None
    assert D.beta_alpha(D.daily_returns([1.0, 2.0]), D.daily_returns([1.0, 2.0]), 1, 1)["beta"] is None


def test_rebased_series_forward_fills_and_never_peeks():
    frame = pd.DataFrame(
        {"Close": [100.0, 110.0]},
        index=pd.to_datetime(["2024-01-01", "2024-01-03"]),
    )
    out = D.rebased_series(frame, [date(2024, 1, 1), date(2024, 1, 2), date(2024, 1, 3)], 1000.0)
    assert out == [1000.0, 1000.0, pytest.approx(1100.0)]


def test_rebased_series_handles_missing_frame():
    assert D.rebased_series(None, [date(2024, 1, 1)], 100.0) == [None]


def test_equal_weight_universe_averages_constituents():
    idx = pd.to_datetime(["2024-01-01", "2024-01-02"])
    frames = {
        "A": pd.DataFrame({"Close": [100.0, 110.0]}, index=idx),   # +10%
        "B": pd.DataFrame({"Close": [100.0, 90.0]}, index=idx),    # -10%
    }
    out = D.equal_weight_universe(frames, [date(2024, 1, 1), date(2024, 1, 2)], 1000.0)
    assert out[1] == pytest.approx(1000.0)  # the two legs cancel


# ── Fill journal ─────────────────────────────────────────────────────────────


def _position(symbol="ACME"):
    return Position(
        symbol=symbol, quantity=10, entry_price=100.0, entry_date=date(2024, 4, 1),
        target_price=120.0, target_pct=20.0, trailing_stop_pct=8.0,
        stop_distance=8.0, stop_price=92.0, highest_price=100.0, sector="Tech",
        method="static",
    )


def test_fills_are_journalled_for_both_legs():
    pf = Portfolio(cash=100_000.0, commission_pct=0.20)
    assert pf.open_position(_position())
    pf.positions["ACME"].highest_price = 130.0
    trade = pf.close_position("ACME", 120.0, date(2024, 6, 1), "target")

    assert [f.side for f in pf.fills] == ["BUY", "SELL"]
    buy, sell = pf.fills
    assert buy.seq == 1 and sell.seq == 2
    assert buy.cost == pytest.approx(1000.0 * 0.002)
    assert sell.anchor == 130.0                     # ratcheted stop anchor
    assert sell.cash_after == pytest.approx(pf.cash)
    assert sell.net_pnl == pytest.approx(trade.pnl)
    assert sell.holding_days == 61
    assert buy.net_pnl is None                       # only sells realise P&L


def test_closed_trade_splits_gross_pnl_from_costs():
    pf = Portfolio(cash=100_000.0, commission_pct=0.20)
    pf.open_position(_position())
    trade = pf.close_position("ACME", 120.0, date(2024, 6, 1), "target")
    assert trade.gross_pnl == pytest.approx(200.0)
    assert trade.costs == pytest.approx(1000.0 * 0.002 + 1200.0 * 0.002)
    assert trade.pnl == pytest.approx(trade.gross_pnl - trade.costs)


def test_rejected_open_is_not_journalled():
    pf = Portfolio(cash=10.0, commission_pct=0.20)
    assert not pf.open_position(_position())
    assert pf.fills == []


# ── Workbook assembly ────────────────────────────────────────────────────────


def _fake_run(days=120):
    """A minimal engine/prices pair good enough to exercise the whole builder."""
    cfg = live_mirror_config(starting_capital=500_000.0)
    pf = Portfolio(cash=500_000.0, commission_pct=0.20)
    pf.open_position(_position())
    pf.close_position("ACME", 120.0, date(2024, 6, 1), "target")

    calendar = pd.bdate_range("2024-04-01", periods=days)
    daily = [
        {"date": d.date().isoformat(), "equity": 500_000.0 + i * 100,
         "cash": 400_000.0, "deployed": 100_000.0 + i * 100, "open_positions": 2}
        for i, d in enumerate(calendar)
    ]
    engine = SimpleNamespace(pf=pf, daily_log=daily)
    bench = pd.DataFrame({"Close": [100.0 + i * 0.1 for i in range(days)]}, index=calendar)
    prices = SimpleNamespace(benchmark=bench, frames={"ACME": bench})
    return cfg, engine, prices


def test_build_dossier_returns_every_reference_sheet():
    cfg, engine, prices = _fake_run()
    sheets = D.build_dossier(cfg, engine, prices, notes=["note"])
    assert set(sheets) == set(D.SHEETS)
    assert list(sheets["Positions"].columns)[:5] == [
        "ticker", "industry", "entry_date", "exit_date", "hold_days"
    ]
    assert list(sheets["Trades"].columns)[-3:] == ["anchor", "stop_level", "cash_after"]
    assert len(sheets["Equity_Curve"]) == 120
    assert len(sheets["Daily_Returns_Portfolio"]) == 119
    assert sheets["Rolling_5Y"].empty          # window is far too short
    assert not sheets["Tax_Ledger"].empty


def test_build_dossier_rejects_an_empty_run():
    cfg, engine, prices = _fake_run()
    engine.daily_log = []
    engine.pf.equity_curve = []
    with pytest.raises(ValueError):
        D.build_dossier(cfg, engine, prices)


def test_net_curve_is_never_above_the_pre_tax_curve():
    cfg, engine, prices = _fake_run()
    sheets = D.build_dossier(cfg, engine, prices)
    net = sheets["Equity_Curve"]["portfolio (net)"].tolist()
    pre_tax = [row["equity"] for row in engine.daily_log]
    assert all(n <= p + 1e-6 for n, p in zip(net, pre_tax))


def test_workbook_writes_all_sheets_with_notes_for_empty_ones(tmp_path):
    cfg, engine, prices = _fake_run()
    sheets = D.build_dossier(cfg, engine, prices, notes=["caveat"])
    out = D.write_workbook(tmp_path / "d.xlsx", sheets,
                           titles={"Positions": "Closed round trips"})
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == list(D.SHEETS)

    # Reference layout: Positions header sits on row 3, Yearly_Returns on row 4.
    assert wb["Positions"].cell(row=3, column=1).value == "ticker"
    assert wb["Yearly_Returns"].cell(row=4, column=1).value == "Year"
    # An impossible sheet explains itself rather than being silently blank.
    assert "5-year" in str(wb["Rolling_5Y"].cell(row=2, column=1).value)


def test_excel_safe_coerces_numpy_and_nan():
    import numpy as np
    assert D._excel_safe(np.int64(3)) == 3
    assert D._excel_safe(np.float64(1.5)) == 1.5
    assert D._excel_safe(float("nan")) is None
    assert D._excel_safe(np.bool_(True)) is True
    assert D._excel_safe(pd.Timestamp("2024-01-01")) == date(2024, 1, 1)
