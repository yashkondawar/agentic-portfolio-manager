"""
Tests for the GFS results dossier.

The point of a dossier is that someone can trust the numbers without rerunning
the backtest, so the tests here are mostly *reconciliation* tests: they assert
that the sheets add up to each other and to the equity curve. A formatting bug
is annoying; a sheet that quietly disagrees with the cash book is worse than no
sheet at all.
"""

from datetime import date

import pandas as pd
import pytest

from backtesting.gfs import dossier as D
from backtesting.gfs.engine import GFSBacktestEngine
from backtesting.gfs.panels import (
    build_panels,
    build_qualify_matrix,
    build_regime_panel,
    build_sector_panel,
    master_calendar,
)
from backtesting.gfs.taxes import TaxConfig
from backtesting.gfs.universe import UniverseStock
from test_gfs_engine import FakeData, build_market, make_cfg, random_walk


def build_bull_market(num_symbols=12, n=2200, seed0=300):
    """A market that trends up.

    The default synthetic market is a driftless random walk, on which GFS loses
    money - so it never generates a tax bill and is useless for exercising the
    settlement path. Tax code that only runs on profitable books needs a
    profitable book to run on.
    """
    idx = pd.bdate_range("2014-01-01", periods=n)
    frames = {
        f"SYM{i:02d}": random_walk(n, seed0 + i, drift=0.0011, index=idx)
        for i in range(num_symbols)
    }
    bench = random_walk(
        n, seed0 + 999, drift=0.0008, vol=0.010, start=10_000.0, index=idx
    )
    sectors = ["IT", "Bank", "Pharma", "Auto"]
    universe = [
        UniverseStock(symbol=s, industry=sectors[i % len(sectors)])
        for i, s in enumerate(sorted(frames))
    ]
    return FakeData(frames, bench), universe


def run_engine(data, universe, cfg, tax_config=None):
    panels = build_panels(data, universe, cfg)
    calendar = master_calendar(data.benchmark, panels)
    sector = build_sector_panel(panels, calendar, cfg)
    regime = build_regime_panel(data.benchmark, panels, calendar, cfg)
    qualify = build_qualify_matrix(panels, calendar, cfg)
    engine = GFSBacktestEngine(
        cfg, panels, sector, regime, qualify, calendar, tax_config=tax_config
    )
    engine.run(cfg.start_date, cfg.end_date)
    return engine, panels, calendar


def _prepared(data, universe):
    """Stands in for ``PreparedData``: the dossier only ever calls ``panels_for``."""

    class Prepared:
        def panels_for(self, cfg):
            panels = build_panels(data, universe, cfg)
            calendar = master_calendar(data.benchmark, panels)
            return (
                panels,
                calendar,
                build_sector_panel(panels, calendar, cfg),
                build_regime_panel(data.benchmark, panels, calendar, cfg),
                build_qualify_matrix(panels, calendar, cfg),
            )

    return Prepared()


@pytest.fixture(scope="module")
def market():
    return build_market()


@pytest.fixture(scope="module")
def bull_market():
    return build_bull_market()


@pytest.fixture(scope="module")
def taxed_run(bull_market):
    data, universe = bull_market
    # Cash yield off: it is the one contribution to equity that no per-trade row
    # accounts for, so switching it off makes the reconciliation identity exact
    # rather than approximate.
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0, cash_yield_pct=0.0)
    engine, panels, calendar = run_engine(data, universe, cfg, TaxConfig())
    return engine, panels, calendar, cfg


# ── the numbers tie out ─────────────────────────────────────────────────────


def test_the_equity_curve_equals_capital_plus_pnl_minus_tax(taxed_run):
    """The single most important property of the whole workbook.

    Final equity must be explainable entirely by rows the reader can see:
    realised P&L (net of both commission legs), tax actually paid, and the
    unrealised mark on whatever was still open. If this drifts, some cost or
    credit is being applied that no sheet discloses.
    """
    engine, _, calendar, cfg = taxed_run
    last_ts = pd.Timestamp(engine.daily_log[-1]["date"])
    price = engine._price_lookup(last_ts)

    realised = sum(t.net_pnl for t in engine.pf.closed)
    tax_paid = sum(x["amount"] for x in engine.pf.taxes_paid)
    unrealised = sum(
        ((price(p.symbol) or p.entry_price) - p.entry_price) * p.quantity - p.entry_cost
        for p in engine.pf.positions.values()
    )
    expected = cfg.starting_capital + realised - tax_paid + unrealised
    assert engine.daily_log[-1]["equity"] == pytest.approx(expected, abs=0.05)


def test_every_rupee_of_assessed_tax_is_actually_paid(taxed_run):
    """The ledger and the cash book must agree.

    The closing financial year's bill only falls due the following April, which
    is after the backtest ends. Without a final settlement the last year's gains
    would be reported in Tax_Ledger but never deducted from equity - inflating
    the final value by exactly the amount the sheet says is owed.
    """
    engine, _, _, _ = taxed_run
    ledger = D._tax_ledger_rows(engine, TaxConfig())
    assessed = sum(r["total tax paid"] for r in ledger)
    paid = sum(x["amount"] for x in engine.pf.taxes_paid)
    assert assessed > 0, "fixture produced no taxable gains - test proves nothing"
    assert paid == pytest.approx(assessed, abs=0.01)


def test_a_positions_row_reconciles_gross_costs_and_net(taxed_run):
    engine, _, _, _ = taxed_run
    price = engine._price_lookup(pd.Timestamp(engine.daily_log[-1]["date"]))
    rows = D._positions_rows(engine, TaxConfig(), price)
    assert rows
    for r in rows:
        assert r["net_pnl"] == pytest.approx(r["gross_pnl"] - r["costs"], abs=0.01)
        # A row is short-term or long-term, never both and never neither.
        if r["status"] != "open":
            assert (r["st_gain"] == 0.0) != (r["lt_gain"] == 0.0) or r["net_pnl"] == 0.0


def test_the_fill_ledger_covers_both_legs_of_every_round_trip(taxed_run):
    """Positions is round-trip shaped; Trades is fill shaped. A scale-out makes
    them differ, so the invariant is a relationship, not equality."""
    engine, _, _, _ = taxed_run
    fills = engine.pf.fills
    buys = [f for f in fills if f["side"] == "BUY"]
    sells = [f for f in fills if f["side"] == "SELL"]
    assert len(sells) == len(engine.pf.closed)
    assert len(buys) == len(engine.pf.closed) + len(engine.pf.positions) - sum(
        1 for t in engine.pf.closed if t.partial
    )
    assert [f["seq"] for f in fills] == list(range(1, len(fills) + 1))


def test_cash_after_on_the_last_fill_is_not_stale(taxed_run):
    """``cash_after`` is only useful if it was captured after the fill moved
    cash, not before."""
    engine, _, _, cfg = taxed_run
    first_buy = next(f for f in engine.pf.fills if f["side"] == "BUY")
    assert first_buy["cash_after"] == pytest.approx(
        cfg.starting_capital - first_buy["value"] - first_buy["cost"], abs=0.01
    )


# ── the tax model behaves ───────────────────────────────────────────────────


def test_tax_is_off_unless_asked_for(market):
    """The live runner and every existing study must be unaffected."""
    data, universe = market
    cfg = make_cfg(s_rsi_entry=45.0)
    engine, _, _ = run_engine(data, universe, cfg, tax_config=None)
    assert engine.pf.taxes_paid == []


def test_paying_tax_lowers_the_final_equity(bull_market):
    data, universe = bull_market
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0, cash_yield_pct=0.0)
    untaxed, _, _ = run_engine(data, universe, cfg, tax_config=None)
    taxed, _, _ = run_engine(data, universe, cfg, tax_config=TaxConfig())
    assert sum(x["amount"] for x in taxed.pf.taxes_paid) > 0
    assert taxed.daily_log[-1]["equity"] < untaxed.daily_log[-1]["equity"]


def test_a_losing_book_is_never_charged_tax(market):
    """Losses carry forward; they do not generate a refund and they must not
    generate a bill either."""
    data, universe = market
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0, cash_yield_pct=0.0)
    engine, _, _ = run_engine(data, universe, cfg, TaxConfig())
    assert sum(t.net_pnl for t in engine.pf.closed) < 0, "fixture is not a losing book"
    assert sum(x["amount"] for x in engine.pf.taxes_paid) == 0.0


def test_a_financial_year_is_never_settled_twice(taxed_run):
    engine, _, _, _ = taxed_run
    labels = [x["financial_year"] for x in engine.pf.taxes_paid]
    assert len(labels) == len(set(labels))


def test_tax_uses_the_costs_the_book_actually_paid(taxed_run):
    """Statutory charges are modelled separately from execution costs, so taxing
    a statutory-charge P&L would tax a number the cash book never saw."""
    engine, _, _, _ = taxed_run
    table = D.tax_mod.apply_to_trades(
        engine.pf.closed, TaxConfig(), use_recorded_costs=True
    )
    for row, trade in zip(table.itertuples(), engine.pf.closed):
        assert row.charges == pytest.approx(trade.total_cost, abs=1e-6)
        assert row.net_pnl == pytest.approx(trade.net_pnl, abs=1e-6)


# ── statistics ──────────────────────────────────────────────────────────────


def test_beta_against_itself_is_one_and_alpha_is_zero():
    rets = [0.01, -0.02, 0.005, 0.011, -0.004, 0.02, -0.011]
    beta, alpha, corr = D._beta_alpha_corr(rets, rets)
    assert beta == pytest.approx(1.0)
    assert corr == pytest.approx(1.0)
    assert alpha == pytest.approx(0.0, abs=1e-9)
    assert D._tracking_error(rets, rets) == pytest.approx(0.0)


def test_a_levered_copy_has_beta_two_and_no_alpha():
    bench = [0.01, -0.02, 0.005, 0.011, -0.004, 0.02, -0.011]
    port = [2 * r for r in bench]
    beta, alpha, corr = D._beta_alpha_corr(port, bench)
    assert beta == pytest.approx(2.0)
    assert corr == pytest.approx(1.0)
    # Twice the mean return minus twice the mean return is zero: leverage alone
    # earns no alpha, which is the whole reason the metric exists.
    assert alpha == pytest.approx(0.0, abs=1e-9)


def test_sortino_ignores_upside_volatility():
    """A series with big up days and small down days should score better on
    Sortino than on Sharpe - otherwise the metric is not doing its job."""
    rets = [0.05, 0.04, -0.002, 0.06, -0.001, 0.05, -0.003]
    assert D._sortino(rets) > D._sharpe(rets)


def test_max_drawdown_is_measured_from_the_running_peak():
    assert D._max_drawdown([100, 120, 60, 200]) == pytest.approx(-0.5)
    assert D._max_drawdown([100, 110, 120]) == pytest.approx(0.0)


def test_calendar_year_return_is_measured_from_the_prior_year_close():
    dates = [pd.Timestamp(d) for d in
             ["2020-06-01", "2020-12-31", "2021-06-01", "2021-12-31"]]
    out = D._calendar_year_returns(dates, [100.0, 110.0, 120.0, 132.0])
    # 2020 is a partial year from inception, so 100 -> 110.
    assert out[2020] == pytest.approx(0.10)
    # 2021 starts from the 2020 close, not from its own first mark.
    assert out[2021] == pytest.approx(0.20)


def test_a_rolling_window_spans_the_requested_number_of_years():
    idx = pd.bdate_range("2015-01-01", periods=1500)
    equity = [100.0 * (1.0002 ** i) for i in range(len(idx))]
    windows = D._rolling_windows(list(idx), equity, 3)
    assert windows
    for start, end, cagr in windows:
        assert 2.9 < (end - start).days / 365.25 < 3.1
        assert cagr > 0


def test_rolling_windows_are_empty_when_the_sample_is_too_short():
    idx = pd.bdate_range("2020-01-01", periods=200)
    assert D._rolling_windows(list(idx), [100.0] * 200, 5) == []


# ── the equal-weight benchmark ──────────────────────────────────────────────


def test_the_equal_weight_basket_tracks_a_single_name_universe(market):
    """With one symbol, the basket must be that symbol - a sanity check that the
    rebasing and compounding are not introducing a drift of their own."""
    data, universe = market
    cfg = make_cfg()
    panels = build_panels(data, universe, cfg)
    one = {"SYM00": panels["SYM00"]}
    calendar = list(panels["SYM00"].frame.index[-200:])
    curve = D._universe_equal_weight(one, calendar, 1000.0)
    closes = panels["SYM00"].frame["Close"].loc[calendar]
    expected = 1000.0 * closes.iloc[-1] / closes.iloc[0]
    assert curve[0] == pytest.approx(1000.0)
    assert curve[-1] == pytest.approx(expected, rel=1e-9)


def test_a_symbol_listing_mid_sample_does_not_create_a_fake_return(market):
    """A name whose first ever print lands inside the window must contribute
    nothing before that day. Treating "no price" -> "price" as a return would
    manufacture an enormous gain out of a listing event."""
    from dataclasses import replace

    data, universe = market
    cfg = make_cfg()
    panels = build_panels(data, universe, cfg)
    early = panels["SYM00"]
    calendar = list(early.frame.index[-200:])
    # SYM01 only starts trading 100 sessions into the window.
    late = replace(panels["SYM01"], frame=panels["SYM01"].frame.loc[calendar[100]:])

    both = D._universe_equal_weight({"SYM00": early, "SYM01": late}, calendar, 1000.0)
    solo = D._universe_equal_weight({"SYM00": early}, calendar, 1000.0)

    # Before the listing the basket is a single name, so it must match exactly.
    assert both[:100] == pytest.approx(solo[:100])
    # And the listing day itself must not print a jump.
    assert abs(both[100] / both[99] - 1.0) < 0.25


# ── the workbook itself ─────────────────────────────────────────────────────


def test_the_workbook_has_every_expected_sheet(tmp_path, market, monkeypatch):
    data, universe = market
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0)

    # The synthetic market has no real index data, so the reference series come
    # back empty. The workbook must still build - a missing benchmark is a
    # degraded report, not a crash.
    monkeypatch.setattr(D, "_index_series", lambda *a, **k: None)
    out = D.build_dossier(cfg, _prepared(data, universe), tmp_path / "d.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert wb.sheetnames == [
        "Summary", "Equity_Curve", "Positions", "Trades", "Yearly_Returns",
        "Rolling_3Y", "Rolling_5Y", "Daily_Returns_Portfolio", "Tax_Ledger",
    ]
    assert wb["Equity_Curve"].max_row > 100


def test_the_empty_sheet_headers_match_the_real_ones(tmp_path, market, monkeypatch):
    """The fallback column lists exist so an empty sheet still has a header.
    If they drift from what the row builders actually emit, an empty Rolling_5Y
    would describe different columns than a populated one."""
    data, universe = market
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0)
    monkeypatch.setattr(D, "_index_series", lambda *a, **k: None)
    out = D.build_dossier(cfg, _prepared(data, universe), tmp_path / "d.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    for name, expected in D._FALLBACK_COLUMNS.items():
        ws = wb[name]
        header = next(
            r for r in range(1, 12)
            if ws.cell(r, 1).value and ws.cell(r, 2).value
        )
        actual = [ws.cell(header, c).value for c in range(1, len(expected) + 1)]
        assert actual == expected, name
        # Nothing beyond the declared columns.
        assert ws.cell(header, len(expected) + 1).value is None, name


def test_data_starts_immediately_under_the_header(tmp_path, market, monkeypatch):
    """No blank spacer between a header and its first data row.

    openpyxl reports ``max_row == 1`` for a still-empty sheet, so predicting the
    header row instead of reading it back silently bolds the first data row and
    leaves a gap - which breaks anyone reading the sheet with a fixed header
    offset.
    """
    data, universe = market
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0)
    monkeypatch.setattr(D, "_index_series", lambda *a, **k: None)
    out = D.build_dossier(cfg, _prepared(data, universe), tmp_path / "d.xlsx")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    for name in wb.sheetnames:
        if name == "Summary":
            continue
        ws = wb[name]
        header = next(
            r for r in range(1, 12)
            if ws.cell(r, 1).value and ws.cell(r, 2).value
        )
        # Rolling_5Y is legitimately empty on a short backtest: header, no rows.
        # What must never happen is a header followed by a blank row followed by
        # data.
        assert ws.max_row == header or ws.cell(header + 1, 1).value is not None, (
            f"{name}: blank row between header and data"
        )


def test_excel_safe_strips_values_openpyxl_cannot_store():
    assert D._excel_safe(float("nan")) is None
    assert D._excel_safe(float("inf")) is None
    assert D._excel_safe(pd.Timestamp("2024-01-02")) == date(2024, 1, 2)
    assert D._excel_safe(None) is None
    assert D._excel_safe(3.5) == 3.5


def test_the_three_friction_variants_are_ordered(market):
    """Removing costs cannot make a strategy worse, and adding tax cannot make
    it better. If this ever fails, the variants are not what they claim."""
    data, universe = market
    cfg = make_cfg(s_rsi_entry=45.0, exit_rsi=65.0, cash_yield_pct=0.0)
    v = D.run_variants(cfg, _prepared(data, universe), TaxConfig())
    assert v["net"].final <= v["pre_tax"].final
    assert v["pre_tax"].final <= v["raw"].final
